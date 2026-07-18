"""
ULU Mahsuri Villa — Operations Accountant (Cloud Version)
Complete Streamlit app for STR income/expense tracking, profit sharing, ROI and LHDN reporting.
Requirements: pip install streamlit anthropic reportlab pypdf pdfplumber pillow pandas supabase
"""

import streamlit as st
import os
import base64
import json
import datetime
import io
from pathlib import Path

import pandas as pd
from PIL import Image
import anthropic
import pdfplumber
from pypdf import PdfReader
from supabase import create_client, Client

from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.lib.units import mm
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, HRFlowable
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.enums import TA_LEFT, TA_RIGHT, TA_CENTER

# ─────────────────────────────────────────────
# SUPABASE CONNECTION
# ─────────────────────────────────────────────
SUPABASE_URL = st.secrets.get("SUPABASE_URL", os.environ.get("SUPABASE_URL",""))
SUPABASE_KEY = st.secrets.get("SUPABASE_KEY", os.environ.get("SUPABASE_KEY",""))

@st.cache_resource
def get_supabase() -> Client:
    return create_client(SUPABASE_URL, SUPABASE_KEY)

# ─────────────────────────────────────────────
# SUPABASE DB ADAPTER
# Mimics sqlite3 connection interface so rest of app stays unchanged
# ─────────────────────────────────────────────
class SupabaseRow(dict):
    """Dict that also supports attribute access like sqlite3.Row"""
    def __getitem__(self, key):
        if isinstance(key, int):
            return list(self.values())[key]
        return super().__getitem__(key)

class SupabaseConn:
    """Drop-in replacement for sqlite3 connection"""
    def __init__(self):
        self.sb = get_supabase()
        self._pending = []  # batch of (table, op, data) for commit

    def execute(self, sql, params=None):
        return SupabaseCursor(self.sb, sql, params)

    def commit(self):
        pass  # Supabase writes are immediate

    def close(self):
        pass  # No connection to close

class SupabaseCursor:
    """Executes SQL-like operations via Supabase REST"""
    def __init__(self, sb, sql, params=None):
        self.sb = sb
        self.sql = sql.strip()
        self.params = params or []
        self._rows = []
        self._execute()

    def _execute(self):
        sql = self.sql
        p   = self.params

        try:
            # ── SELECT ──────────────────────────────────────────────────────
            if sql.upper().startswith("SELECT"):
                self._rows = self._handle_select(sql, p)

            # ── INSERT ──────────────────────────────────────────────────────
            elif sql.upper().startswith("INSERT"):
                self._handle_insert(sql, p)

            # ── UPDATE ──────────────────────────────────────────────────────
            elif sql.upper().startswith("UPDATE"):
                self._handle_update(sql, p)

            # ── DELETE ──────────────────────────────────────────────────────
            elif sql.upper().startswith("DELETE"):
                self._handle_delete(sql, p)

        except Exception as e:
            # Re-raise so callers can catch and surface the real error.
            # Only show st.error for SELECT failures (non-fatal reads).
            if sql.upper().startswith("SELECT"):
                st.error(f"DB read error: {e}\nSQL: {sql}\nParams: {p}")
            else:
                raise  # Let INSERT/UPDATE/DELETE callers handle and display

    def _table_from_sql(self, sql):
        import re
        m = re.search(r'FROM\s+(\w+)', sql, re.IGNORECASE)
        if not m:
            m = re.search(r'INTO\s+(\w+)', sql, re.IGNORECASE)
        if not m:
            m = re.search(r'UPDATE\s+(\w+)', sql, re.IGNORECASE)
        if not m:
            m = re.search(r'DELETE FROM\s+(\w+)', sql, re.IGNORECASE)
        return m.group(1) if m else None

    def _handle_select(self, sql, p):
        import re
        table = self._table_from_sql(sql)
        if not table:
            return []

        # Build query
        q = self.sb.table(table).select("*")

        # Parse WHERE conditions
        where_match = re.search(r'WHERE\s+(.+?)(?:ORDER|LIMIT|GROUP|$)', sql, re.IGNORECASE | re.DOTALL)
        if where_match:
            where_clause = where_match.group(1).strip()
            q = self._apply_where(q, table, where_clause, p)

        # ORDER BY
        order_match = re.search(r'ORDER BY\s+(.+?)(?:LIMIT|$)', sql, re.IGNORECASE)
        if order_match:
            order_str = order_match.group(1).strip()
            parts = order_str.split(',')
            for part in parts:
                part = part.strip()
                desc = 'DESC' in part.upper()
                col = re.sub(r'\s+(ASC|DESC)', '', part, flags=re.IGNORECASE).strip()
                q = q.order(col, desc=desc)

        # LIMIT
        limit_match = re.search(r'LIMIT\s+(\d+)', sql, re.IGNORECASE)
        if limit_match:
            q = q.limit(int(limit_match.group(1)))
        else:
            q = q.limit(10000)

        # COUNT(*)
        if re.search(r'COUNT\(\*\)', sql, re.IGNORECASE) and 'FROM' in sql.upper():
            result = q.execute()
            count_val = len(result.data) if result.data else 0
            # Return as single row with count
            self._rows = [SupabaseRow({"COUNT(*)": count_val, "t": count_val})]
            return self._rows

        # SUM / COALESCE SUM
        sum_match = re.search(r'(?:COALESCE\()?SUM\((\w+)\)(?:,\s*\d+\))?\s+(?:as\s+)?(\w+)?', sql, re.IGNORECASE)
        if sum_match:
            col_name = sum_match.group(1)
            alias    = sum_match.group(2) or "t"
            result = q.execute()
            total = sum(float(r.get(col_name) or 0) for r in (result.data or []))
            self._rows = [SupabaseRow({alias: total, col_name: total, 0: total})]
            return self._rows

        result = q.execute()
        self._rows = [SupabaseRow(r) for r in (result.data or [])]
        return self._rows

    def _apply_where(self, q, table, where_clause, p):
        import re
        # Replace ? with actual values
        idx = [0]
        def next_param():
            v = p[idx[0]] if idx[0] < len(p) else None
            idx[0] += 1
            return v

        # Split on AND
        conditions = re.split(r'\s+AND\s+', where_clause, flags=re.IGNORECASE)
        for cond in conditions:
            cond = cond.strip()
            # Handle IN (?,?,?)
            in_match = re.match(r'(\w+)\s+IN\s*\(([^)]+)\)', cond, re.IGNORECASE)
            if in_match:
                col = in_match.group(1)
                placeholders = in_match.group(2)
                count = placeholders.count('?')
                vals = [next_param() for _ in range(count)]
                q = q.in_(col, vals)
                continue
            # Handle LIKE
            like_match = re.match(r'(\w+)\s+LIKE\s+\?', cond, re.IGNORECASE)
            if like_match:
                col = like_match.group(1)
                val = next_param()
                if val:
                    val = val.replace('%','')
                    q = q.ilike(col, f'%{val}%')
                continue
            # Handle = ? and != ?
            eq_match = re.match(r'(\w+)\s*(=|!=|<|>|<=|>=)\s*\?', cond, re.IGNORECASE)
            if eq_match:
                col, op = eq_match.group(1), eq_match.group(2)
                val = next_param()
                if op == '=':   q = q.eq(col, val)
                elif op == '!=': q = q.neq(col, val)
                elif op == '>':  q = q.gt(col, val)
                elif op == '<':  q = q.lt(col, val)
                elif op == '>=': q = q.gte(col, val)
                elif op == '<=': q = q.lte(col, val)
                continue
            # Handle col != 0 (used for delete all)
            neq_match = re.match(r'(\w+)\s*!=\s*(\d+)', cond, re.IGNORECASE)
            if neq_match:
                col, val = neq_match.group(1), int(neq_match.group(2))
                q = q.neq(col, val)
        return q

    def _handle_insert(self, sql, p):
        import re
        table = self._table_from_sql(sql)
        if not table:
            return
        # Extract column names
        col_match = re.search(r'\(([^)]+)\)\s+VALUES', sql, re.IGNORECASE)
        if not col_match:
            return
        cols = [c.strip() for c in col_match.group(1).split(',')]
        record = {}
        for i, col in enumerate(cols):
            record[col] = p[i] if i < len(p) else None
        self.sb.table(table).insert(record).execute()

    def _handle_update(self, sql, p):
        import re
        table = self._table_from_sql(sql)
        if not table:
            return
        set_match   = re.search(r'SET\s+(.+?)\s+WHERE', sql, re.IGNORECASE | re.DOTALL)
        where_match = re.search(r'WHERE\s+(.+?)$',      sql, re.IGNORECASE | re.DOTALL)
        if not set_match:
            return
        set_clause   = set_match.group(1).strip()
        where_clause = where_match.group(1).strip() if where_match else ""

        # ── Parse SET clause carefully ────────────────────────────────────────
        # Each assignment is either:
        #   col = ?           → consume next positional param
        #   col = 'literal'   → use the literal value directly (no param consumed)
        # A naive len(set_cols) count breaks when literals are present.
        record = {}
        param_idx = 0
        for assignment in set_clause.split(','):
            assignment = assignment.strip()
            # col = 'literal string'
            lit_match = re.match(r"(\w+)\s*=\s*'([^']*)'", assignment)
            if lit_match:
                record[lit_match.group(1)] = lit_match.group(2)
                continue
            # col = NULL
            null_match = re.match(r"(\w+)\s*=\s*NULL", assignment, re.IGNORECASE)
            if null_match:
                record[null_match.group(1)] = None
                continue
            # col = ?  → consume next param
            placeholder_match = re.match(r"(\w+)\s*=\s*\?", assignment)
            if placeholder_match:
                col = placeholder_match.group(1)
                record[col] = p[param_idx] if param_idx < len(p) else None
                param_idx += 1
                continue

        where_vals = list(p[param_idx:])

        q = self.sb.table(table).update(record)

        # Apply WHERE
        if where_vals and 'id' in where_clause.lower():
            q = q.eq('id', where_vals[0])
        elif where_vals and where_clause:
            conditions = re.split(r'\s+AND\s+', where_clause, flags=re.IGNORECASE)
            widx = 0
            for cond in conditions:
                col_match = re.match(r'(\w+)\s*=\s*\?', cond.strip(), re.IGNORECASE)
                if col_match and widx < len(where_vals):
                    q = q.eq(col_match.group(1), where_vals[widx])
                    widx += 1
        q.execute()

    def _handle_delete(self, sql, p):
        import re
        table = self._table_from_sql(sql)
        if not table:
            return
        where_match = re.search(r'WHERE\s+(.+?)$', sql, re.IGNORECASE | re.DOTALL)
        q = self.sb.table(table).delete()
        if where_match and p:
            where_clause = where_match.group(1).strip()
            conditions = re.split(r'\s+AND\s+', where_clause, flags=re.IGNORECASE)
            param_idx = 0
            for cond in conditions:
                cond = cond.strip()
                eq_match = re.match(r'(\w+)\s*=\s*\?', cond, re.IGNORECASE)
                neq_match= re.match(r'(\w+)\s*!=\s*\?', cond, re.IGNORECASE)
                if eq_match and param_idx < len(p):
                    q = q.eq(eq_match.group(1), p[param_idx]); param_idx += 1
                elif neq_match and param_idx < len(p):
                    q = q.neq(neq_match.group(1), p[param_idx]); param_idx += 1
        q.execute()

    def fetchone(self):
        return self._rows[0] if self._rows else None

    def fetchall(self):
        return self._rows

    def __iter__(self):
        return iter(self._rows)


def get_db():
    return SupabaseConn()

def init_db():
    pass  # Tables already created in Supabase via SQL Editor

DB_PATH = "supabase"

# ─────────────────────────────────────────────
# PAGE CONFIG
# ─────────────────────────────────────────────
st.set_page_config(
    page_title="ULU Accountant",
    page_icon="🌾",
    layout="wide",
    initial_sidebar_state="expanded",
)

# ─────────────────────────────────────────────
# CUSTOM CSS
# ─────────────────────────────────────────────
st.markdown("""
<style>
  @import url('https://fonts.googleapis.com/css2?family=DM+Serif+Display&family=DM+Sans:wght@300;400;500;600&display=swap');
  html, body, [class*="css"] { font-family: 'DM Sans', sans-serif; }
  h1,h2,h3 { font-family: 'DM Serif Display', serif; }

  .stApp { background-color: #F5F0E8; }

  /* ── Header: ID-anchored so no global span rule can override it ── */
  #ulu-header, #ulu-header * {
    color: #F5F0E8 !important;
  }
  #ulu-header {
    background: linear-gradient(135deg, #2a3528 0%, #1a2418 100%);
    padding: 28px 36px 22px 36px;
    border-radius: 12px;
    margin-bottom: 24px;
  }
  #ulu-header h1 {
    font-family:'DM Serif Display',serif;
    font-size:2rem; margin:0 0 4px 0;
    color:#F5F0E8 !important;
  }
  #ulu-header p {
    margin:0; color:#A8B8A5 !important;
    font-size:0.85rem; letter-spacing:0.05em; text-transform:uppercase;
  }

  /* ── Global contrast & font-size fixes ── */
  /* Body text, markdown, labels */
  [class*="css"] p, [class*="css"] li,
  .stMarkdown p, .stMarkdown li,
  label, .stTextInput label, .stSelectbox label,
  .stNumberInput label, .stTextArea label,
  .stDateInput label, .stFileUploader label {
    color: #1C1C1A !important;
    font-size: 1rem !important;
  }
  /* Input/textarea field values */
  input, textarea,
  .stTextInput input, .stTextArea textarea,
  .stNumberInput input {
    color: #1C1C1A !important;
    font-size: 1rem !important;
  }
  /* Selectbox displayed value */
  [data-baseweb="select"] [data-baseweb="tag"],
  [data-baseweb="select"] div[class*="singleValue"],
  [data-baseweb="select"] div[class*="placeholder"] {
    color: #1C1C1A !important;
    font-size: 1rem !important;
  }
  /* Dropdown portal menu items — rendered outside normal DOM */
  [data-baseweb="popover"] li,
  [data-baseweb="menu"] li,
  [data-baseweb="list-item"],
  ul[data-baseweb="menu"] span {
    color: #1C1C1A !important;
    font-size: 1rem !important;
  }

  .card {
    background:#FFFFFF;
    border-radius:10px;
    padding:22px 26px;
    margin-bottom:16px;
    border:1px solid #E5DDD0;
  }
  .card-title {
    font-family:'DM Serif Display',serif;
    font-size:1.1rem;
    color:#1C1C1A;
    margin:0 0 14px 0;
  }

  .metric-box {
    background:#2a3528;
    color:#F5F0E8;
    border-radius:10px;
    padding:18px 20px;
    text-align:center;
    margin-bottom:8px;
  }
  .metric-box.accent { background:#C4856A; }
  .metric-box.mid    { background:#4a5e47; }
  .metric-label { font-size:0.75rem; letter-spacing:0.06em; text-transform:uppercase; color:#A8B8A5; margin-bottom:5px; }
  .metric-box.accent .metric-label { color:#f0d5c8; }
  .metric-box.mid    .metric-label { color:#c5d4c3; }
  .metric-value { font-family:'DM Serif Display',serif; font-size:1.75rem; line-height:1; }

  .stTabs [data-baseweb="tab-list"] {
    background:#FFFFFF; border-radius:10px; padding:5px; gap:3px;
    border:1px solid #E5DDD0; margin-bottom:18px;
  }
  .stTabs [data-baseweb="tab"] {
    border-radius:7px; padding:8px 16px;
    font-weight:500; font-size:0.85rem; color:#6B6560;
  }
  .stTabs [aria-selected="true"] { background:#2a3528 !important; color:#F5F0E8 !important; }

  .stButton > button {
    background:#2a3528; color:#F5F0E8; border:none;
    border-radius:8px; padding:10px 22px;
    font-weight:500; font-size:0.88rem; transition:opacity 0.15s;
  }
  .stButton > button:hover { opacity:0.82; }

  .badge-whole { background:#2a3528; color:white; padding:2px 10px; border-radius:12px; font-size:0.75rem; }
  .badge-mbed  { background:#C4856A; color:white; padding:2px 10px; border-radius:12px; font-size:0.75rem; }
  .badge-airbnb  { background:#FF5A5F; color:white; padding:2px 8px; border-radius:10px; font-size:0.72rem; }
  .badge-direct  { background:#2a3528; color:white; padding:2px 8px; border-radius:10px; font-size:0.72rem; }

  /* Payment voucher status badges */
  .badge-paid     { background:#2a7a2a; color:white; padding:2px 10px; border-radius:12px; font-size:0.8rem; font-weight:600; }
  .badge-pending  { background:#b8860b; color:white; padding:2px 10px; border-radius:12px; font-size:0.8rem; font-weight:600; }
  .badge-partial  { background:#4a90d9; color:white; padding:2px 10px; border-radius:12px; font-size:0.8rem; font-weight:600; }
  .badge-cancelled{ background:#555; color:white; padding:2px 10px; border-radius:12px; font-size:0.8rem; font-weight:600; }
</style>
""", unsafe_allow_html=True)

# ─────────────────────────────────────────────
# HELPERS
# ─────────────────────────────────────────────

# ─────────────────────────────────────────────
# HELPERS
# ─────────────────────────────────────────────
MONTHS = ["Jan","Feb","Mar","Apr","May","Jun","Jul","Aug","Sep","Oct","Nov","Dec"]
EXPENSE_CATEGORIES = [
    "Pool Maintenance","WiFi","Water Filter (Coway)","Electricity (TNB)",
    "Water (SADA)","Housekeeping & Laundry","Operation Items","Supplies & Toiletries",
    "FF&E (Furniture & Equipment)","Maintenance & Repairs","Marketing","Insurance",
    "Assessment / Quit Rent","Miscellaneous"
]

# OpEx classification for manager expense lines
OPEX_CATEGORIES = [
    "Uncategorised",
    "Utilities — Electricity (TNB)",
    "Utilities — Water (SADA)",
    "Utilities — WiFi (UNIFI)",
    "Utilities — Water Filter (Coway)",
    "Housekeeping & Laundry",
    "Pool Maintenance",
    "Maintenance & Minor Repairs",
    "Operation Items & Supplies",
    "Marketing",
    "Insurance",
    "Assessment / Quit Rent",
    "Property Management Fee (Archmedia)",
    "Miscellaneous OpEx",
]

# CapEx categories for upgrade/improvement works
CAPEX_CATEGORIES = [
    "Furniture & Furnishings (FF&E)",
    "Soft Furnishings & Decor",
    "Kitchen Equipment & Appliances",
    "Bathroom Fittings & Fixtures",
    "Electrical & Lighting",
    "Renovation & Structural Works",
    "Outdoor & Landscaping",
    "Technology & Smart Home",
    "Pool & Recreation Equipment",
    "Miscellaneous CapEx",
]

# OpEx vs CapEx classification rule (for reference)
CAPEX_THRESHOLD = 1000  # Items above RM1,000 with multi-year life = CapEx

def get_setting(key):
    conn = get_db()
    r = conn.execute("SELECT value FROM settings WHERE key=?", (key,)).fetchone()
    conn.close()
    return r["value"] if r else None

def set_setting(key, value):
    conn = get_db()
    conn.execute("INSERT OR REPLACE INTO settings VALUES (?,?)", (key, str(value)))
    conn.commit()
    conn.close()

def get_year_month_list():
    """Return list of (year, month) tuples from operation start to now."""
    start_y = int(get_setting("operation_start_year") or 2024)
    start_m = int(get_setting("operation_start_month") or 7)
    now = datetime.datetime.now()
    result = []
    y, m = start_y, start_m
    while (y, m) <= (now.year, now.month):
        result.append((y, m))
        m += 1
        if m > 12:
            m = 1
            y += 1
    return list(reversed(result))

def operation_year(year, month):
    """Calculate Y01/Y02 etc based on July 2024 start."""
    start_y = int(get_setting("operation_start_year") or 2024)
    start_m = int(get_setting("operation_start_month") or 7)
    total_months = (year - start_y) * 12 + (month - start_m)
    op_year = (total_months // 12) + 1
    return f"Y{op_year:02d}"

def fmt_myr(v):
    return f"RM {float(v or 0):,.2f}"

def get_monthly_summary(year, month):
    conn = get_db()
    gross_income = conn.execute(
        "SELECT COALESCE(SUM(amount),0) as t FROM bookings WHERE year=? AND month=?",
        (year, month)
    ).fetchone()["t"]

    mgr_total = conn.execute(
        "SELECT COALESCE(SUM(amount),0) as t FROM manager_expenses WHERE year=? AND month=?",
        (year, month)
    ).fetchone()["t"]

    personal_total = conn.execute(
        "SELECT COALESCE(SUM(ulu_share),0) as t FROM personal_expenses WHERE year=? AND month=?",
        (year, month)
    ).fetchone()["t"]

    conn.close()

    gross_op_cost = mgr_total + personal_total
    net_before_sharing = gross_income - gross_op_cost
    cohost_pct = float(get_setting("cohost_pct") or 30)
    cohost_share = max(0, net_before_sharing * cohost_pct / 100)
    owner_share = max(0, net_before_sharing - cohost_share)

    return {
        "gross_income": gross_income,
        "mgr_expenses": mgr_total,
        "personal_expenses": personal_total,
        "gross_op_cost": gross_op_cost,
        "net_before_sharing": net_before_sharing,
        "cohost_pct": cohost_pct,
        "cohost_share": cohost_share,
        "owner_share": owner_share,
    }

def get_yearly_summary(year):
    # Get all months that have bookings or expenses — avoid GROUP BY which breaks Supabase adapter
    conn = get_db()
    bk_rows  = conn.execute("SELECT DISTINCT month FROM bookings WHERE year=?", (year,)).fetchall()
    ex_rows  = conn.execute("SELECT DISTINCT month FROM manager_expenses WHERE year=?", (year,)).fetchall()
    pe_rows  = conn.execute("SELECT DISTINCT month FROM personal_expenses WHERE year=?", (year,)).fetchall()
    conn.close()

    all_months = set()
    for r in bk_rows:  all_months.add(int(r["month"]))
    for r in ex_rows:  all_months.add(int(r["month"]))
    for r in pe_rows:  all_months.add(int(r["month"]))

    monthly = {}
    for month in sorted(all_months):
        monthly[month] = get_monthly_summary(year, month)
    return monthly

# ─────────────────────────────────────────────
# AI EXTRACTION
# ─────────────────────────────────────────────
def extract_receipt(file_bytes, file_name, api_key):
    client_ai = anthropic.Anthropic(api_key=api_key)
    ext = Path(file_name).suffix.lower()

    if ext == ".pdf":
        text_content = ""
        try:
            with pdfplumber.open(io.BytesIO(file_bytes)) as pdf:
                for page in pdf.pages:
                    text_content += (page.extract_text() or "")
        except Exception:
            pass
        if not text_content.strip():
            try:
                reader = PdfReader(io.BytesIO(file_bytes))
                for page in reader.pages:
                    text_content += (page.extract_text() or "")
            except Exception:
                pass

        prompt = f"""You are an expert accountant in Malaysia reviewing an expense receipt for a villa business.
Extract these fields from the receipt text:

{text_content[:3000]}

Reply ONLY with valid JSON (no markdown):
{{
  "vendor": "vendor/shop name",
  "bill_date": "YYYY-MM-DD",
  "description": "1 sentence summary of what was purchased",
  "total_amount": 0.00,
  "suggested_category": "one of: Pool Maintenance, WiFi, Water Filter (Coway), Electricity (TNB), Water (SADA), Housekeeping & Laundry, Operation Items, Supplies & Toiletries, FF&E (Furniture & Equipment), Maintenance & Repairs, Marketing, Insurance, Miscellaneous"
}}

Rules: total_amount is a plain number, no currency symbols."""

        response = client_ai.messages.create(
            model="claude-sonnet-4-6", max_tokens=400,
            messages=[{"role": "user", "content": prompt}]
        )
        raw = response.content[0].text.strip().replace("```json","").replace("```","").strip()
        return json.loads(raw)

    else:
        media_type = "image/jpeg" if ext in [".jpg",".jpeg"] else "image/png"
        b64 = base64.standard_b64encode(file_bytes).decode("utf-8")
        prompt = """You are an expert accountant in Malaysia reviewing an expense receipt for a villa business.
Extract these fields from this receipt image.

Reply ONLY with valid JSON (no markdown):
{
  "vendor": "vendor/shop name",
  "bill_date": "YYYY-MM-DD",
  "description": "1 sentence summary of what was purchased",
  "total_amount": 0.00,
  "suggested_category": "one of: Pool Maintenance, WiFi, Water Filter (Coway), Electricity (TNB), Water (SADA), Housekeeping & Laundry, Operation Items, Supplies & Toiletries, FF&E (Furniture & Equipment), Maintenance & Repairs, Marketing, Insurance, Miscellaneous"
}

Rules: total_amount is a plain number, no RM symbol."""

        response = client_ai.messages.create(
            model="claude-sonnet-4-6", max_tokens=400,
            messages=[{"role": "user", "content": [
                {"type": "image", "source": {"type": "base64", "media_type": media_type, "data": b64}},
                {"type": "text", "text": prompt}
            ]}]
        )
        raw = response.content[0].text.strip().replace("```json","").replace("```","").strip()
        return json.loads(raw)

# ─────────────────────────────────────────────
# AI EXTRACTION — AZARY'S MONTHLY BILLING REPORT
# ─────────────────────────────────────────────
def extract_manager_report(file_bytes, file_name, api_key):
    """
    Extract full structured data from Azary's monthly billing report.
    Returns dict with bookings list, expenses list, and summary figures.
    """
    import anthropic as _ant
    client_ai = _ant.Anthropic(api_key=api_key, timeout=90.0)
    ext = Path(file_name).suffix.lower()

    prompt = """You are extracting data from a ULU Mahsuri Villa Monthly Billing Report.

This document has 4 sections. You MUST read and extract ALL of them:

SECTION 1 — INCOME SUMMARY (top table): Contains guest bookings with columns: NO, GUEST NAME, VILLA, CHECK-IN, CHECK-OUT, NIGHT, SOURCE, AMOUNT (RM), NOTES. Extract every row.

SECTION 2 — OPERATING EXPENSES: Contains expense lines with columns: No, Expense Item, Vendor/Payee, Pay By, Account No., Bank, Amount (RM), Notes. Extract every row.

SECTION 3 — NET PROFIT CALCULATION: Contains summary rows: Gross Income, Operating Expenses, Net Profit Before Sharing, Profit Sharing Owner (70%), Profit Sharing Co-Host (30%).

SECTION 4 — PAYMENT RECORD: Ignore this section.

Also read the document header for: MONTH, YEAR, OCCUPANCY %.

Return ONLY valid JSON (no markdown, no preamble, no explanation):

{
  "month": 4,
  "year": 2026,
  "occupancy_pct": 79,
  "bookings": [
    {
      "guest_name": "OUMAIMA OUAISSA",
      "room_type": "MBED",
      "checkin": "2026-04-04",
      "checkout": "2026-04-07",
      "nights": 3,
      "source": "AIRBNB",
      "amount": 1644.92,
      "notes": ""
    }
  ],
  "expenses": [
    {
      "expense_item": "Pool Cleaner",
      "vendor": "Qasim Bin Ismail",
      "amount": 200.00,
      "notes": ""
    }
  ],
  "gross_income": 13570.46,
  "total_opex": 4092.67,
  "net_profit": 9477.79,
  "owner_share": 6634.45,
  "cohost_share": 2843.34
}

Strict rules:
- room_type: "MBED" if VILLA column shows M-BED, M-Bed, MASTER BED, MBED. "WHOLE" if it shows WHOLE, Whole.
- source: "AIRBNB" if source is Airbnb/AIRBNB. "DIRECT" for anything else.
- checkin/checkout: YYYY-MM-DD format. Use the month/year from the header to fill in the year.
- All amounts are plain numbers, no RM symbol, no commas.
- Include ALL booking rows from Section 1, including CANCELLED bookings.
- Include ALL expense rows from Section 2.
- gross_income = TOTAL SALE from Section 1 = sum of all booking amounts.
- total_opex = TOTAL from Section 2 = sum of all expense amounts.
- Return ONLY the JSON object. Nothing before or after it."""

    if ext == ".pdf":
        # Try text extraction first
        text_content = ""
        try:
            with pdfplumber.open(io.BytesIO(file_bytes)) as pdf:
                for page in pdf.pages:
                    text_content += (page.extract_text() or "")
        except Exception:
            pass
        if not text_content.strip():
            try:
                reader = PdfReader(io.BytesIO(file_bytes))
                for page in reader.pages:
                    text_content += (page.extract_text() or "")
            except Exception:
                pass

        if text_content.strip():
            # Use text mode
            full_prompt = f"{prompt}\n\nDocument text:\n{text_content[:6000]}"
            response = client_ai.messages.create(
                model="claude-sonnet-4-6", max_tokens=2000,
                messages=[{"role": "user", "content": full_prompt}]
            )
        else:
            # Fall back to vision on PDF base64
            b64 = base64.standard_b64encode(file_bytes).decode("utf-8")
            response = client_ai.messages.create(
                model="claude-sonnet-4-6", max_tokens=2000,
                messages=[{"role": "user", "content": [
                    {"type": "document", "source": {"type": "base64", "media_type": "application/pdf", "data": b64}},
                    {"type": "text", "text": prompt}
                ]}]
            )
    else:
        # Image
        media_type = "image/jpeg" if ext in [".jpg", ".jpeg"] else "image/png"
        b64 = base64.standard_b64encode(file_bytes).decode("utf-8")
        response = client_ai.messages.create(
            model="claude-sonnet-4-6", max_tokens=2000,
            messages=[{"role": "user", "content": [
                {"type": "image", "source": {"type": "base64", "media_type": media_type, "data": b64}},
                {"type": "text", "text": prompt}
            ]}]
        )

    raw = response.content[0].text.strip()
    raw = raw.replace("```json", "").replace("```", "").strip()
    return json.loads(raw)


def save_manager_scan_file(file_bytes, file_name, year, month):
    """Cloud version — no local disk save. Return placeholder path."""
    return f"cloud/{year}-{month:02d}/{file_name}"


# ─────────────────────────────────────────────
# PDF MONTHLY REPORT
# ─────────────────────────────────────────────
def generate_monthly_report(year, month, summary, bookings, mgr_expenses, personal_expenses):
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4,
                            rightMargin=18*mm, leftMargin=18*mm,
                            topMargin=16*mm, bottomMargin=16*mm)

    INK    = colors.HexColor("#1C1C1A")
    GREEN  = colors.HexColor("#2a3528")
    CREAM  = colors.HexColor("#F5F0E8")
    ACCENT = colors.HexColor("#C4856A")
    LIGHT  = colors.HexColor("#E5DDD0")
    GREY   = colors.HexColor("#6B6560")

    s_normal = ParagraphStyle("n", fontName="Helvetica", fontSize=9, leading=13, textColor=INK)
    s_small  = ParagraphStyle("s", fontName="Helvetica", fontSize=8, leading=12, textColor=GREY)
    s_bold   = ParagraphStyle("b", fontName="Helvetica-Bold", fontSize=9, leading=13, textColor=INK)
    s_title  = ParagraphStyle("t", fontName="Helvetica-Bold", fontSize=16, leading=20, textColor=INK)
    s_sub    = ParagraphStyle("su", fontName="Helvetica-Bold", fontSize=11, leading=14, textColor=GREEN)
    s_right  = ParagraphStyle("r", fontName="Helvetica", fontSize=9, alignment=TA_RIGHT, textColor=INK)
    s_right_b= ParagraphStyle("rb",fontName="Helvetica-Bold", fontSize=10, alignment=TA_RIGHT, textColor=ACCENT)
    s_label  = ParagraphStyle("l", fontName="Helvetica", fontSize=7.5, textColor=GREY, spaceAfter=1)

    W = A4[0] - 36*mm
    story = []

    # Header
    op_yr = operation_year(year, month)
    story.append(Paragraph(f"ULU Mahsuri Villa", s_title))
    story.append(Paragraph(f"Monthly Operations Report — {MONTHS[month-1]} {year} ({op_yr})", s_sub))
    story.append(HRFlowable(width="100%", thickness=2, color=GREEN, spaceAfter=14))

    # Summary boxes as table
    def box_row(label, value, highlight=False):
        lp = ParagraphStyle("bl", fontName="Helvetica", fontSize=8.5,
                            textColor=GREY if not highlight else colors.white)
        vp = ParagraphStyle("bv", fontName="Helvetica-Bold", fontSize=11,
                            alignment=TA_RIGHT,
                            textColor=ACCENT if highlight else INK)
        return [Paragraph(label, lp), Paragraph(value, vp)]

    summary_data = [
        box_row("Gross Income", fmt_myr(summary["gross_income"])),
        box_row("Manager Expenses", f"({fmt_myr(summary['mgr_expenses'])})"),
        box_row("Personal Expenses (ULU Share)", f"({fmt_myr(summary['personal_expenses'])})"),
        box_row("Gross Operating Cost", f"({fmt_myr(summary['gross_op_cost'])})"),
        box_row("Net Profit Before Sharing", fmt_myr(summary["net_before_sharing"])),
        box_row(f"Co-Host Share ({summary['cohost_pct']:.0f}%)", f"({fmt_myr(summary['cohost_share'])})"),
        box_row("Owner Net Profit (Your Income)", fmt_myr(summary["owner_share"]), highlight=True),
    ]
    sum_tbl = Table(summary_data, colWidths=[W*0.65, W*0.35])
    sum_tbl.setStyle(TableStyle([
        ("LINEBELOW", (0,4),(1,4), 1, GREEN),
        ("LINEABOVE", (0,6),(1,6), 1, INK),
        ("TOPPADDING",(0,0),(-1,-1),5),
        ("BOTTOMPADDING",(0,0),(-1,-1),5),
        ("LEFTPADDING",(0,0),(-1,-1),6),
        ("RIGHTPADDING",(0,0),(-1,-1),0),
        ("BACKGROUND",(0,6),(1,6), colors.HexColor("#F5F0E8")),
    ]))
    story.append(sum_tbl)
    story.append(Spacer(1, 18))

    # Bookings table
    story.append(Paragraph("1. Booking Income", s_sub))
    story.append(Spacer(1, 6))
    if bookings:
        bk_head = [
            Paragraph("Guest", ParagraphStyle("th", fontName="Helvetica-Bold", fontSize=8, textColor=colors.white)),
            Paragraph("Type", ParagraphStyle("th", fontName="Helvetica-Bold", fontSize=8, textColor=colors.white)),
            Paragraph("Check-in", ParagraphStyle("th", fontName="Helvetica-Bold", fontSize=8, textColor=colors.white)),
            Paragraph("Check-out", ParagraphStyle("th", fontName="Helvetica-Bold", fontSize=8, textColor=colors.white)),
            Paragraph("Nights", ParagraphStyle("th", fontName="Helvetica-Bold", fontSize=8, textColor=colors.white)),
            Paragraph("Source", ParagraphStyle("th", fontName="Helvetica-Bold", fontSize=8, textColor=colors.white)),
            Paragraph("Amount", ParagraphStyle("th", fontName="Helvetica-Bold", fontSize=8, alignment=TA_RIGHT, textColor=colors.white)),
        ]
        bk_rows = [bk_head]
        for b in bookings:
            bk_rows.append([
                Paragraph(b["guest_name"] or "", s_small),
                Paragraph(b["room_type"] or "", s_small),
                Paragraph(b["checkin"] or "", s_small),
                Paragraph(b["checkout"] or "", s_small),
                Paragraph(str(b["nights"] or ""), s_small),
                Paragraph(b["source"] or "", s_small),
                Paragraph(f"{float(b['amount'] or 0):,.2f}", ParagraphStyle("ra", fontName="Helvetica", fontSize=8, alignment=TA_RIGHT, textColor=INK)),
            ])
        bk_tbl = Table(bk_rows, colWidths=[W*0.18, W*0.08, W*0.12, W*0.12, W*0.07, W*0.1, W*0.13])
        bk_tbl.setStyle(TableStyle([
            ("BACKGROUND",(0,0),(-1,0), GREEN),
            ("ROWBACKGROUNDS",(0,1),(-1,-1),[colors.white, colors.HexColor("#FAF8F5")]),
            ("TOPPADDING",(0,0),(-1,-1),5),("BOTTOMPADDING",(0,0),(-1,-1),5),
            ("LEFTPADDING",(0,0),(-1,-1),6),("RIGHTPADDING",(0,0),(-1,-1),6),
            ("LINEBELOW",(0,0),(-1,-1),0.3, LIGHT),
        ]))
        story.append(bk_tbl)
    else:
        story.append(Paragraph("No bookings recorded.", s_small))
    story.append(Spacer(1, 14))

    # Manager expenses
    story.append(Paragraph("2. Manager's Monthly Expenses", s_sub))
    story.append(Spacer(1, 6))
    if mgr_expenses:
        me_head = [
            Paragraph("Item", ParagraphStyle("th", fontName="Helvetica-Bold", fontSize=8, textColor=colors.white)),
            Paragraph("Vendor / Payee", ParagraphStyle("th", fontName="Helvetica-Bold", fontSize=8, textColor=colors.white)),
            Paragraph("Amount", ParagraphStyle("th", fontName="Helvetica-Bold", fontSize=8, alignment=TA_RIGHT, textColor=colors.white)),
        ]
        me_rows = [me_head]
        for e in mgr_expenses:
            me_rows.append([
                Paragraph(e["expense_item"] or "", s_small),
                Paragraph(e["vendor"] or "", s_small),
                Paragraph(f"{float(e['amount'] or 0):,.2f}", ParagraphStyle("ra", fontName="Helvetica", fontSize=8, alignment=TA_RIGHT, textColor=INK)),
            ])
        me_tbl = Table(me_rows, colWidths=[W*0.35, W*0.40, W*0.25])
        me_tbl.setStyle(TableStyle([
            ("BACKGROUND",(0,0),(-1,0), GREEN),
            ("ROWBACKGROUNDS",(0,1),(-1,-1),[colors.white, colors.HexColor("#FAF8F5")]),
            ("TOPPADDING",(0,0),(-1,-1),5),("BOTTOMPADDING",(0,0),(-1,-1),5),
            ("LEFTPADDING",(0,0),(-1,-1),6),("RIGHTPADDING",(0,0),(-1,-1),6),
            ("LINEBELOW",(0,0),(-1,-1),0.3, LIGHT),
        ]))
        story.append(me_tbl)
    else:
        story.append(Paragraph("No manager expenses recorded.", s_small))
    story.append(Spacer(1, 14))

    # Personal expenses
    story.append(Paragraph("3. Owner's Personal Expenses (ULU Share)", s_sub))
    story.append(Spacer(1, 6))
    if personal_expenses:
        pe_head = [
            Paragraph("Vendor", ParagraphStyle("th", fontName="Helvetica-Bold", fontSize=8, textColor=colors.white)),
            Paragraph("Date", ParagraphStyle("th", fontName="Helvetica-Bold", fontSize=8, textColor=colors.white)),
            Paragraph("Category", ParagraphStyle("th", fontName="Helvetica-Bold", fontSize=8, textColor=colors.white)),
            Paragraph("Description", ParagraphStyle("th", fontName="Helvetica-Bold", fontSize=8, textColor=colors.white)),
            Paragraph("Total Bill", ParagraphStyle("th", fontName="Helvetica-Bold", fontSize=8, alignment=TA_RIGHT, textColor=colors.white)),
            Paragraph("ULU Share", ParagraphStyle("th", fontName="Helvetica-Bold", fontSize=8, alignment=TA_RIGHT, textColor=colors.white)),
        ]
        pe_rows = [pe_head]
        for e in personal_expenses:
            pe_rows.append([
                Paragraph(e["vendor"] or "", s_small),
                Paragraph(e["bill_date"] or "", s_small),
                Paragraph(e["category"] or "", s_small),
                Paragraph(e["description"] or "", s_small),
                Paragraph(f"{float(e['total_amount'] or 0):,.2f}", ParagraphStyle("ra", fontName="Helvetica", fontSize=8, alignment=TA_RIGHT, textColor=GREY)),
                Paragraph(f"{float(e['ulu_share'] or 0):,.2f}", ParagraphStyle("ra", fontName="Helvetica-Bold", fontSize=8, alignment=TA_RIGHT, textColor=INK)),
            ])
        pe_tbl = Table(pe_rows, colWidths=[W*0.16, W*0.11, W*0.18, W*0.25, W*0.13, W*0.13])
        pe_tbl.setStyle(TableStyle([
            ("BACKGROUND",(0,0),(-1,0), GREEN),
            ("ROWBACKGROUNDS",(0,1),(-1,-1),[colors.white, colors.HexColor("#FAF8F5")]),
            ("TOPPADDING",(0,0),(-1,-1),5),("BOTTOMPADDING",(0,0),(-1,-1),5),
            ("LEFTPADDING",(0,0),(-1,-1),6),("RIGHTPADDING",(0,0),(-1,-1),6),
            ("LINEBELOW",(0,0),(-1,-1),0.3, LIGHT),
        ]))
        story.append(pe_tbl)
    else:
        story.append(Paragraph("No personal expenses recorded.", s_small))

    story.append(Spacer(1, 18))
    story.append(HRFlowable(width="100%", thickness=0.5, color=LIGHT, spaceAfter=8))
    story.append(Paragraph(
        f"Generated on {datetime.datetime.now().strftime('%d %B %Y %H:%M')} · ULU Mahsuri Villa Operations Accountant",
        s_small
    ))

    doc.build(story)
    return buffer.getvalue()



# ─────────────────────────────────────────────
# HEADER
# ─────────────────────────────────────────────
st.markdown("""
<div id="ulu-header">
  <h1 style="color:#F5F0E8 !important;">🌾 ULU Mahsuri Villa</h1>
  <p style="color:#A8B8A5 !important;">Operations Accountant · Langkawi · Personal Income Tracking</p>
</div>
""", unsafe_allow_html=True)

# ─────────────────────────────────────────────
# SIDEBAR — API KEY + SETTINGS
# ─────────────────────────────────────────────
with st.sidebar:
    st.markdown("### ⚙️ Settings")

    # Load from .env if not yet in session
    if not st.session_state.get("ulu_api_key"):
        env_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), ".env")
        if os.path.exists(env_path):
            with open(env_path) as _ef:
                for line in _ef:
                    if "ANTHROPIC_API_KEY" in line:
                        st.session_state["ulu_api_key"] = line.split("=",1)[1].strip().strip('"')

    api_key_input = st.text_input(
        "Claude API Key (for AI extraction)",
        value=st.session_state.get("ulu_api_key",""),
        type="password",
        placeholder="sk-ant-api03-...",
        key="sidebar_api_key",
        help="Required for AI document extraction in CapEx Tracker and Scan Receipts"
    )
    if api_key_input:
        st.session_state["ulu_api_key"] = api_key_input

    st.divider()
    st.caption(f"DB: {DB_PATH}")

# ─────────────────────────────────────────────
# TABS
# ─────────────────────────────────────────────
tab1, tab2, tab3, tab4, tab5, tab6, tab7, tab8, tab9, tab10 = st.tabs([
    "📅  Monthly Entry",
    "🧾  Scan Receipts",
    "📊  Monthly P&L",
    "📈  Yearly Dashboard",
    "🗂️  All Records",
    "🏗️  CapEx Tracker",
    "📋  OpEx Breakdown",
    "📦  Accountant",
    "💳  Payments & Vouchers",
    "💰  Direct Income & Extras",
])

# ══════════════════════════════════════════════
# TAB 1 — MONTHLY ENTRY
# ══════════════════════════════════════════════
with tab1:
    ym_list = get_year_month_list()
    ym_labels = [f"{MONTHS[m-1]} {y} ({operation_year(y,m)})" for y,m in ym_list]
    sel_idx = st.selectbox("Select Month", range(len(ym_labels)), format_func=lambda i: ym_labels[i], key="entry_month")
    sel_year, sel_month = ym_list[sel_idx]

    col_left, col_right = st.columns([1,1], gap="large")

    # ── BOOKINGS ──
    with col_left:
        st.markdown('<div class="card">', unsafe_allow_html=True)
        st.markdown('<p class="card-title">➕ Add Booking</p>', unsafe_allow_html=True)
        with st.form("add_booking", clear_on_submit=True):
            b_guest  = st.text_input("Guest Name", placeholder="e.g. Ahmad Ibrahim")
            b_type   = st.selectbox("Room Type", ["WHOLE", "MBED"],
                                    help="WHOLE = both rooms, MBED = Master Bedroom only")
            b_col1, b_col2 = st.columns(2)
            b_checkin  = b_col1.text_input("Check-in (YYYY-MM-DD)", value=f"{sel_year}-{sel_month:02d}-01")
            b_checkout = b_col2.text_input("Check-out (YYYY-MM-DD)", value=f"{sel_year}-{sel_month:02d}-02")
            b_nights = st.number_input("Nights", min_value=1, step=1, value=1)
            b_source = st.selectbox("Source", ["AIRBNB","DIRECT"])
            b_amount = st.number_input("Amount Received (RM)", min_value=0.0, step=0.01, format="%.2f")
            b_notes  = st.text_input("Notes (optional)")
            if st.form_submit_button("💾 Save Booking", use_container_width=True):
                conn = get_db()
                conn.execute(
                    "INSERT INTO bookings (year,month,guest_name,room_type,checkin,checkout,nights,source,amount,notes) VALUES (?,?,?,?,?,?,?,?,?,?)",
                    (sel_year, sel_month, b_guest, b_type, b_checkin, b_checkout, b_nights, b_source, b_amount, b_notes)
                )
                conn.commit(); conn.close()
                st.success("Booking saved!")
                st.rerun()
        st.markdown('</div>', unsafe_allow_html=True)

        # Show bookings for month
        conn = get_db()
        bks = conn.execute("SELECT * FROM bookings WHERE year=? AND month=? ORDER BY checkin", (sel_year, sel_month)).fetchall()
        conn.close()
        if bks:
            st.markdown('<div class="card">', unsafe_allow_html=True)
            st.markdown(f'<p class="card-title">Bookings — {MONTHS[sel_month-1]} {sel_year}</p>', unsafe_allow_html=True)
            df = pd.DataFrame([dict(b) for b in bks])[["id","guest_name","room_type","checkin","checkout","nights","source","amount"]]
            df.columns = ["ID","Guest","Type","Check-in","Check-out","Nights","Source","Amount (RM)"]
            df["Amount (RM)"] = df["Amount (RM)"].apply(lambda x: f"{float(x):,.2f}")
            st.dataframe(df, use_container_width=True, hide_index=True)
            del_b = st.number_input("Delete booking by ID", min_value=0, step=1, value=0, key="del_b")
            if st.button("🗑️ Delete Booking"):
                if del_b > 0:
                    conn = get_db()
                    conn.execute("DELETE FROM bookings WHERE id=?", (del_b,))
                    conn.commit(); conn.close()
                    st.rerun()
            st.markdown('</div>', unsafe_allow_html=True)

    # ── MANAGER EXPENSES ──
    with col_right:
        st.markdown('<div class="card">', unsafe_allow_html=True)
        st.markdown('<p class="card-title">➕ Manager\'s Monthly Expenses</p>', unsafe_allow_html=True)
        st.caption("Key in line by line from Azary's monthly submission PDF")

        MANAGER_ITEMS = [
            "Pool Cleaner","WiFi (Unifi)","Water Filter (Coway)",
            "Electricity (TNB)","Water (SADA)",
            "Housekeeping & Laundry","Operation Items","Other"
        ]
        with st.form("add_mgr_expense", clear_on_submit=True):
            me_item   = st.selectbox("Expense Item", MANAGER_ITEMS)
            me_vendor = st.text_input("Vendor / Payee", placeholder="e.g. Qasim Bin Ismail")
            me_amount = st.number_input("Amount (RM)", min_value=0.0, step=0.01, format="%.2f")
            me_notes  = st.text_input("Notes (optional)")
            if st.form_submit_button("💾 Save Expense Line", use_container_width=True):
                conn = get_db()
                conn.execute(
                    "INSERT INTO manager_expenses (year,month,expense_item,vendor,amount,notes) VALUES (?,?,?,?,?,?)",
                    (sel_year, sel_month, me_item, me_vendor, me_amount, me_notes)
                )
                conn.commit(); conn.close()
                st.success("Saved!")
                st.rerun()
        st.markdown('</div>', unsafe_allow_html=True)

        # Show manager expenses
        conn = get_db()
        mes = conn.execute("SELECT * FROM manager_expenses WHERE year=? AND month=? ORDER BY id", (sel_year, sel_month)).fetchall()
        conn.close()
        if mes:
            st.markdown('<div class="card">', unsafe_allow_html=True)
            st.markdown('<p class="card-title">Manager Expenses Entered</p>', unsafe_allow_html=True)
            df_me = pd.DataFrame([dict(m) for m in mes])[["id","expense_item","vendor","amount"]]
            df_me.columns = ["ID","Item","Vendor","Amount (RM)"]
            df_me["Amount (RM)"] = df_me["Amount (RM)"].apply(lambda x: f"{float(x):,.2f}")
            st.dataframe(df_me, use_container_width=True, hide_index=True)
            mgr_total = sum(float(m["amount"] or 0) for m in mes)
            st.markdown(f"**Total: {fmt_myr(mgr_total)}**")
            del_me = st.number_input("Delete by ID", min_value=0, step=1, value=0, key="del_me")
            if st.button("🗑️ Delete Expense Line"):
                if del_me > 0:
                    conn = get_db()
                    conn.execute("DELETE FROM manager_expenses WHERE id=?", (del_me,))
                    conn.commit(); conn.close()
                    st.rerun()
            st.markdown('</div>', unsafe_allow_html=True)

# ══════════════════════════════════════════════
# TAB 2 — SCAN RECEIPTS
# ══════════════════════════════════════════════
with tab2:
    scan_subtab1, scan_subtab2, scan_subtab3 = st.tabs([
        "📋 Manager's Monthly Report",
        "📊 Airbnb CSV & Reconcile",
        "🧾 Personal Receipts (Owner)",
    ])

    # ── SUB-TAB A: MANAGER'S MONTHLY REPORT ────────────────────────────────────
    with scan_subtab1:
        st.markdown("**Upload Manager's monthly billing report** — AI extracts all bookings and expenses automatically.")

        api_key_s = st.session_state.get("ulu_api_key", "")

        # Month/year selector at the top — same as Personal Receipts tab
        ym_list_mgr = get_year_month_list()
        ym_labels_mgr = [f"{MONTHS[m-1]} {y} ({operation_year(y,m)})" for y,m in ym_list_mgr]
        sel_idx_mgr = st.selectbox(
            "Assign report to month",
            range(len(ym_labels_mgr)),
            format_func=lambda i: ym_labels_mgr[i],
            key="mgr_report_month"
        )
        mgr_sel_year, mgr_sel_month = ym_list_mgr[sel_idx_mgr]

        col_up, col_prev = st.columns([1, 1], gap="large")

        with col_up:
            st.markdown('<div class="card">', unsafe_allow_html=True)
            st.markdown('<p class="card-title">📤 Upload Monthly Report</p>', unsafe_allow_html=True)
            mgr_upload = st.file_uploader(
                "Drop Manager's monthly billing report (image or PDF)",
                type=["jpg", "jpeg", "png", "pdf"], key="mgr_report_upload"
            )
            if mgr_upload:
                mgr_bytes = mgr_upload.read()
                ext_m = Path(mgr_upload.name).suffix.lower()
                if ext_m in [".jpg", ".jpeg", ".png"]:
                    st.image(Image.open(io.BytesIO(mgr_bytes)), use_container_width=True)
                else:
                    st.info(f"📄 PDF: {mgr_upload.name}")

                if not api_key_s:
                    st.warning("Enter your Anthropic API key in the sidebar.")
                else:
                    if st.button("🤖 Extract Full Report with AI", use_container_width=True, type="primary", key="btn_extract_mgr"):
                        with st.spinner("AI reading manager's report — extracting all bookings and expenses…"):
                            try:
                                result = extract_manager_report(mgr_bytes, mgr_upload.name, api_key_s)
                                st.session_state["mgr_extracted"] = result
                                st.session_state["mgr_file_bytes"] = mgr_bytes
                                st.session_state["mgr_file_name"] = mgr_upload.name
                                st.success(f"✓ Extracted {len(result.get('bookings',[]))} bookings and {len(result.get('expenses',[]))} expense lines.")
                            except Exception as e:
                                st.error(f"Extraction failed: {e}")
            st.markdown('</div>', unsafe_allow_html=True)

        with col_prev:
            if "mgr_extracted" in st.session_state:
                r = st.session_state["mgr_extracted"]
                st.markdown('<div class="card">', unsafe_allow_html=True)
                st.markdown('<p class="card-title">📋 Extracted Data — Review Before Saving</p>', unsafe_allow_html=True)

                # Use selected month from top selector as default; AI-read month as fallback
                ex_year  = r.get("year", mgr_sel_year) or mgr_sel_year
                ex_month = r.get("month", mgr_sel_month) or mgr_sel_month
                # Override with top selector if they match year (trust user selection)
                ex_year  = mgr_sel_year
                ex_month = mgr_sel_month
                ex_occ   = r.get("occupancy_pct", 0)
                st.markdown(f"**Month:** {MONTHS[ex_month-1]} {ex_year} &nbsp;|&nbsp; **Occupancy:** {ex_occ}%")

                # Financials summary
                fc1, fc2, fc3 = st.columns(3)
                fc1.metric("Gross Income", fmt_myr(r.get("gross_income", 0)))
                fc2.metric("Total OpEx", fmt_myr(r.get("total_opex", 0)))
                fc3.metric("Net Profit", fmt_myr(r.get("net_profit", 0)))
                fc4, fc5 = st.columns(2)
                fc4.metric("Owner Share (70%)", fmt_myr(r.get("owner_share", 0)))
                fc5.metric("Co-Host Share (30%)", fmt_myr(r.get("cohost_share", 0)))

                st.markdown("**Bookings extracted:**")
                bk_preview = []
                for b in r.get("bookings", []):
                    bk_preview.append({
                        "Guest": b.get("guest_name",""),
                        "Type": b.get("room_type",""),
                        "Check-in": b.get("checkin",""),
                        "Check-out": b.get("checkout",""),
                        "Nights": b.get("nights",""),
                        "Source": b.get("source",""),
                        "Amount (RM)": f"{float(b.get('amount',0)):,.2f}"
                    })
                if bk_preview:
                    st.dataframe(pd.DataFrame(bk_preview), use_container_width=True, hide_index=True)

                st.markdown("**Expenses extracted:**")
                ex_preview = []
                for e in r.get("expenses", []):
                    ex_preview.append({
                        "Item": e.get("expense_item",""),
                        "Vendor": e.get("vendor",""),
                        "Amount (RM)": f"{float(e.get('amount',0)):,.2f}"
                    })
                if ex_preview:
                    st.dataframe(pd.DataFrame(ex_preview), use_container_width=True, hide_index=True)

                st.markdown('</div>', unsafe_allow_html=True)

                # Confirm & Save
                st.markdown('<div class="card">', unsafe_allow_html=True)
                st.markdown('<p class="card-title">✅ Confirm & Save to Database</p>', unsafe_allow_html=True)

                st.markdown("**Confirm month/year (change if needed):**")
                ov_col1, ov_col2 = st.columns(2)
                override_month = ov_col1.selectbox(
                    "Month", list(range(1,13)),
                    index=mgr_sel_month - 1,
                    format_func=lambda m: MONTHS[m-1],
                    key="mgr_override_month"
                )
                override_year = ov_col2.number_input(
                    "Year", min_value=2024, max_value=2030,
                    value=mgr_sel_year, step=1, key="mgr_override_year"
                )

                st.warning("⚠️ This will add all bookings and expenses to the database. Double-check the month/year above before saving.")

                if st.button("💾 Save All to Database", type="primary", use_container_width=True, key="btn_save_mgr"):
                    try:
                        save_year  = int(override_year)
                        save_month = int(override_month)
                        conn = get_db()
                        saved_bk = 0
                        saved_ex = 0

                        for b in r.get("bookings", []):
                            conn.execute(
                                "INSERT INTO bookings (year,month,guest_name,room_type,checkin,checkout,nights,source,amount,notes) VALUES (?,?,?,?,?,?,?,?,?,?)",
                                (save_year, save_month,
                                 b.get("guest_name",""), b.get("room_type","WHOLE"),
                                 b.get("checkin",""), b.get("checkout",""),
                                 int(b.get("nights",1)), b.get("source","AIRBNB"),
                                 float(b.get("amount",0)), b.get("notes",""))
                            )
                            saved_bk += 1

                        for e in r.get("expenses", []):
                            conn.execute(
                                "INSERT INTO manager_expenses (year,month,expense_item,vendor,amount,notes) VALUES (?,?,?,?,?,?)",
                                (save_year, save_month,
                                 e.get("expense_item",""), e.get("vendor",""),
                                 float(e.get("amount",0)), e.get("notes",""))
                            )
                            saved_ex += 1

                        # Save scan record
                        scan_path = save_manager_scan_file(
                            st.session_state["mgr_file_bytes"],
                            st.session_state["mgr_file_name"],
                            save_year, save_month
                        )
                        conn.execute(
                            """INSERT INTO manager_monthly_scans
                               (year,month,file_name,scan_path,occupancy_pct,gross_income,total_opex,net_profit,owner_share,cohost_share)
                               VALUES (?,?,?,?,?,?,?,?,?,?)""",
                            (save_year, save_month,
                             st.session_state["mgr_file_name"], scan_path,
                             float(r.get("occupancy_pct",0)),
                             float(r.get("gross_income",0)), float(r.get("total_opex",0)),
                             float(r.get("net_profit",0)), float(r.get("owner_share",0)),
                             float(r.get("cohost_share",0)))
                        )
                        conn.commit(); conn.close()

                        for k in ["mgr_extracted","mgr_file_bytes","mgr_file_name"]:
                            st.session_state.pop(k, None)

                        st.success(f"✓ Saved {saved_bk} bookings and {saved_ex} expense lines for {MONTHS[save_month-1]} {save_year}.")
                        st.rerun()
                    except Exception as e:
                        st.error(f"Save failed: {e}")

                if st.button("🗑️ Discard & Start Over", key="btn_discard_mgr"):
                    for k in ["mgr_extracted","mgr_file_bytes","mgr_file_name"]:
                        st.session_state.pop(k, None)
                    st.rerun()
                st.markdown('</div>', unsafe_allow_html=True)
            else:
                st.markdown(
                    "<div style='text-align:center;padding:60px 0;color:#A89F91;'>"
                    "<div style='font-size:3rem'>📋</div>"
                    "<p>Upload Azary's monthly billing report on the left.<br>"
                    "AI will extract all bookings and expenses for your review before saving.</p>"
                    "</div>", unsafe_allow_html=True
                )

        # Previous scans archive
        st.markdown("---")
        st.markdown("#### 📁 Scan Archive — Previously Uploaded Reports")
        st.caption("Delete a scan record to remove it AND all associated bookings and expenses for that month.")
        conn = get_db()
        prev_scans = conn.execute(
            "SELECT * FROM manager_monthly_scans ORDER BY year DESC, month DESC"
        ).fetchall()
        conn.close()
        if prev_scans:
            for s in prev_scans:
                with st.expander(f"📋 {MONTHS[s['month']-1]} {s['year']} — Gross: {fmt_myr(s['gross_income'])} | Owner Net: {fmt_myr(s['owner_share'])} | Occ: {s['occupancy_pct']:.0f}%"):
                    sc1, sc2, sc3, sc4 = st.columns(4)
                    sc1.metric("Gross Income", fmt_myr(s['gross_income']))
                    sc2.metric("Total OpEx", fmt_myr(s['total_opex']))
                    sc3.metric("Owner Share", fmt_myr(s['owner_share']))
                    sc4.metric("Co-Host Share", fmt_myr(s['cohost_share']))

                    st.caption(f"File: {s['file_name']} | Scanned: {s['created_at']}")

                    dl_col, del_col = st.columns([3, 1])
                    sp = s["scan_path"]
                    if sp and os.path.exists(sp):
                        with open(sp, "rb") as _f:
                            ext_sp = sp.split(".")[-1].lower()
                            mime_sp = "application/pdf" if ext_sp == "pdf" else f"image/{ext_sp}"
                            dl_col.download_button(
                                f"⬇️ Download Original Report",
                                data=_f.read(), file_name=os.path.basename(sp), mime=mime_sp,
                                key=f"mgr_dl_{s['id']}"
                            )
                    else:
                        dl_col.caption("Original file not found on disk.")

                    # Delete button with confirmation
                    if del_col.button("🗑️ Delete", key=f"mgr_del_{s['id']}", type="secondary"):
                        st.session_state[f"confirm_del_scan_{s['id']}"] = True

                    if st.session_state.get(f"confirm_del_scan_{s['id']}"):
                        st.error(f"⚠️ This will delete the scan record AND all bookings and manager expenses for **{MONTHS[s['month']-1]} {s['year']}**. Are you sure?")
                        conf1, conf2 = st.columns(2)
                        if conf1.button("✅ Yes, delete everything", key=f"mgr_del_confirm_{s['id']}", type="primary"):
                            conn = get_db()
                            conn.execute("DELETE FROM bookings WHERE year=? AND month=?", (s['year'], s['month']))
                            conn.execute("DELETE FROM manager_expenses WHERE year=? AND month=?", (s['year'], s['month']))
                            conn.execute("DELETE FROM manager_monthly_scans WHERE id=?", (s['id'],))
                            conn.commit(); conn.close()
                            st.session_state.pop(f"confirm_del_scan_{s['id']}", None)
                            st.success(f"✓ Deleted all data for {MONTHS[s['month']-1]} {s['year']}.")
                            st.rerun()
                        if conf2.button("❌ Cancel", key=f"mgr_del_cancel_{s['id']}"):
                            st.session_state.pop(f"confirm_del_scan_{s['id']}", None)
                            st.rerun()
        else:
            st.info("No monthly reports scanned yet.")

    # ── SUB-TAB B: AIRBNB CSV & RECONCILE ──────────────────────────────────────
    with scan_subtab2:
        st.markdown("**Upload the Airbnb monthly CSV** — app compares against the Manager's report already scanned and identifies Direct bookings and Extra charges automatically.")

        ym_list_ab = get_year_month_list()
        ym_labels_ab = [f"{MONTHS[m-1]} {y} ({operation_year(y,m)})" for y,m in ym_list_ab]
        sel_idx_ab = st.selectbox(
            "Month to reconcile",
            range(len(ym_labels_ab)),
            format_func=lambda i: ym_labels_ab[i],
            key="ab_month"
        )
        ab_year, ab_month = ym_list_ab[sel_idx_ab]

        ab_upload = st.file_uploader(
            "Upload Airbnb CSV for this month",
            type=["csv"], key="ab_csv_upload"
        )
        if ab_upload:
            ab_bytes = ab_upload.read()
            st.session_state["ab_csv_bytes"] = ab_bytes
            st.session_state["ab_csv_name"]  = ab_upload.name
            st.success(f"✓ {ab_upload.name} ready")

        if st.session_state.get("ab_csv_bytes") and st.button(
            "🔍 Compare Airbnb vs Manager Report", type="primary", key="btn_ab_reconcile"
        ):
            with st.spinner("Reconciling..."):
                try:
                    airbnb_rows = parse_airbnb_csv(st.session_state["ab_csv_bytes"])
                    conn = get_db()
                    db_bk = conn.execute(
                        "SELECT * FROM bookings WHERE year=? AND month=? ORDER BY checkin",
                        (ab_year, ab_month)
                    ).fetchall()
                    conn.close()
                    db_bookings = [dict(b) for b in db_bk]

                    if not db_bookings:
                        st.warning(f"No Manager report found for {MONTHS[ab_month-1]} {ab_year}. "
                                   f"Please scan the Manager's Monthly Report first (sub-tab above).")
                    else:
                        result = reconcile_airbnb_vs_manager(airbnb_rows, db_bookings)
                        st.session_state["ab_reconcile_result"] = result
                        st.session_state["ab_reconcile_year"]   = ab_year
                        st.session_state["ab_reconcile_month"]  = ab_month
                        if result["directs"]:
                            st.success(f"✓ Done — {len(result['directs'])} Direct/Extra entries identified.")
                        else:
                            st.success("✓ Done — all bookings matched. No Direct/Extras this month.")
                except Exception as e:
                    st.error(f"Reconciliation failed: {e}")

        # Show results
        if "ab_reconcile_result" in st.session_state:
            res  = st.session_state["ab_reconcile_result"]
            r_yr = st.session_state["ab_reconcile_year"]
            r_mo = st.session_state["ab_reconcile_month"]

            st.markdown(f"### {MONTHS[r_mo-1]} {r_yr} — Reconciliation Result")

            m1, m2, m3, m4 = st.columns(4)
            m1.metric("Manager Total", fmt_myr(res["db_total"]))
            m2.metric("Airbnb Payout", fmt_myr(res["airbnb_total"]))
            m3.metric("Direct/Extras", fmt_myr(res["direct_total"]))
            m4.metric("Difference",    fmt_myr(res["difference"]))

            # Matched
            st.markdown("---")
            st.markdown("**✅ Airbnb Bookings — Matched:**")
            if res["matched"]:
                st.dataframe(pd.DataFrame([{
                    "Guest":         m["guest"],
                    "Check-in":      m["checkin"],
                    "Check-out":     m["checkout"],
                    "Nights":        m["nights"],
                    "Airbnb Payout": fmt_myr(m["airbnb_payout"]),
                    "Airbnb Gross":  fmt_myr(m["airbnb_gross"]),
                    "DB Amount":     fmt_myr(m["db_amount"]),
                    "Diff":          fmt_myr(m["amount_diff"]) if abs(m["amount_diff"]) > 0.10 else "✅",
                    "Confirmation":  m["confirmation"],
                } for m in res["matched"]]), use_container_width=True, hide_index=True)

            # Directs
            if res["directs"]:
                st.markdown("---")
                st.markdown("**⚡ Direct Bookings & Extras — Not in Airbnb CSV:**")
                st.warning(f"{len(res['directs'])} booking(s) in Manager report not found in Airbnb — classify below then save.")

                direct_entries = []
                for i, d in enumerate(res["directs"]):
                    with st.expander(
                        f"**{d['guest']}** — {fmt_myr(d['amount'])} — {d['checkin']} to {d['checkout']}",
                        expanded=True
                    ):
                        dc1, dc2 = st.columns(2)
                        d_type   = dc1.selectbox("Income Type", DIRECT_INCOME_TYPES, key=f"ab_type_{i}")
                        d_method = dc2.selectbox("Payment Method",
                            ["Cash","Bank Transfer","DuitNow","Other"], key=f"ab_method_{i}")
                        d_ref    = st.text_input("Reference / Remarks", key=f"ab_ref_{i}",
                            placeholder="e.g. Cash on arrival, bank transfer ref")
                        d_notes  = st.text_input("Notes", key=f"ab_notes_{i}")
                        direct_entries.append({
                            "guest":          d["guest"],
                            "amount":         d["amount"],
                            "checkin":        d["checkin"],
                            "income_type":    d_type,
                            "payment_method": d_method,
                            "reference":      d_ref,
                            "notes":          d_notes,
                        })

                st.markdown("---")
                if st.button("💾 Save Direct/Extra Entries to Register", type="primary", key="btn_ab_save"):
                    conn = get_db()
                    saved = 0
                    errors = []
                    for entry in direct_entries:
                        try:
                            conn.execute(
                                """INSERT INTO direct_income
                                   (year,month,guest_name,income_type,amount,date_received,
                                    payment_method,reference,airbnb_booking_ref,notes)
                                   VALUES (?,?,?,?,?,?,?,?,?,?)""",
                                (r_yr, r_mo, entry["guest"], entry["income_type"],
                                 entry["amount"], entry["checkin"],
                                 entry["payment_method"], entry["reference"], "", entry["notes"])
                            )
                            conn.commit()
                            saved += 1
                        except Exception as e:
                            errors.append(f"{entry['guest']}: {e}")
                    conn.close()
                    if errors:
                        for err in errors: st.error(f"❌ {err}")
                    else:
                        st.success(f"✓ {saved} entries saved to Direct Income register (Tab 10).")
                        for k in ["ab_reconcile_result","ab_reconcile_year",
                                  "ab_reconcile_month","ab_csv_bytes","ab_csv_name"]:
                            st.session_state.pop(k, None)
                        st.rerun()

                if st.button("🗑️ Discard", key="btn_ab_discard"):
                    for k in ["ab_reconcile_result","ab_reconcile_year",
                              "ab_reconcile_month","ab_csv_bytes","ab_csv_name"]:
                        st.session_state.pop(k, None)
                    st.rerun()
            else:
                st.success("✅ All bookings matched — no Direct/Extras this month.")
                if st.button("✅ Done", key="btn_ab_done"):
                    for k in ["ab_reconcile_result","ab_reconcile_year",
                              "ab_reconcile_month","ab_csv_bytes","ab_csv_name"]:
                        st.session_state.pop(k, None)
                    st.rerun()

    # ── SUB-TAB C: PERSONAL RECEIPTS (OWNER) ───────────────────────────────────
    with scan_subtab3:
        st.markdown("**Scan your personal one-off receipts** — AI reads them automatically. Enter ULU's allocated share.")

        ym_list2 = get_year_month_list()
        ym_labels2 = [f"{MONTHS[m-1]} {y} ({operation_year(y,m)})" for y,m in ym_list2]
        sel_idx2 = st.selectbox("Assign to month", range(len(ym_labels2)), format_func=lambda i: ym_labels2[i], key="scan_month")
        sel_year2, sel_month2 = ym_list2[sel_idx2]

        col_scan, col_result = st.columns([1, 1], gap="large")

        with col_scan:
            st.markdown('<div class="card">', unsafe_allow_html=True)
            st.markdown('<p class="card-title">Upload Receipt</p>', unsafe_allow_html=True)
            uploaded = st.file_uploader("Drag & drop receipt", type=["jpg","jpeg","png","pdf"], key="personal_receipt_upload")

            if uploaded:
                file_bytes_p = uploaded.read()
                ext_p = Path(uploaded.name).suffix.lower()
                if ext_p in [".jpg",".jpeg",".png"]:
                    st.image(Image.open(io.BytesIO(file_bytes_p)), use_container_width=True)
                else:
                    st.info(f"📄 PDF: {uploaded.name}")

                if not api_key_s:
                    st.warning("Enter your Anthropic API key in the sidebar.")
                else:
                    if st.button("🤖 Extract with AI", use_container_width=True, key="btn_extract_personal"):
                        with st.spinner("Reading receipt…"):
                            try:
                                extracted = extract_receipt(file_bytes_p, uploaded.name, api_key_s)
                                st.session_state["ulu_extracted"] = extracted
                                st.session_state["ulu_file_name"] = uploaded.name
                                st.session_state["ulu_year"] = sel_year2
                                st.session_state["ulu_month"] = sel_month2
                                st.success("Done! Review on the right.")
                            except Exception as e:
                                st.error(f"Error: {e}")
            st.markdown('</div>', unsafe_allow_html=True)

            with st.expander("➕ Add manually (without AI)"):
                with st.form("manual_personal", clear_on_submit=True):
                    mp_vendor = st.text_input("Vendor")
                    mp_date   = st.text_input("Date", value=datetime.date.today().isoformat())
                    mp_cat    = st.selectbox("Category", EXPENSE_CATEGORIES)
                    mp_desc   = st.text_area("Description", height=60)
                    mp_total  = st.number_input("Total Bill Amount (RM)", min_value=0.0, step=0.01, format="%.2f")
                    mp_share  = st.number_input("ULU's Share (RM)", min_value=0.0, step=0.01, format="%.2f",
                                                help="Enter the portion allocated to ULU Mahsuri Villa only")
                    if st.form_submit_button("💾 Save", use_container_width=True):
                        conn = get_db()
                        conn.execute(
                            "INSERT INTO personal_expenses (year,month,vendor,bill_date,category,description,total_amount,ulu_share,file_name) VALUES (?,?,?,?,?,?,?,?,?)",
                            (sel_year2, sel_month2, mp_vendor, mp_date, mp_cat, mp_desc, mp_total, mp_share, "manual")
                        )
                        conn.commit(); conn.close()
                        st.success("Saved!")
                        st.rerun()

        with col_result:
            st.markdown('<div class="card">', unsafe_allow_html=True)
            st.markdown('<p class="card-title">Review & Confirm</p>', unsafe_allow_html=True)

            if "ulu_extracted" in st.session_state:
                ext_data = st.session_state["ulu_extracted"]
                with st.form("confirm_personal"):
                    r_vendor  = st.text_input("Vendor", value=ext_data.get("vendor",""))
                    r_date    = st.text_input("Date", value=ext_data.get("bill_date",""))
                    r_cat     = st.selectbox("Category", EXPENSE_CATEGORIES,
                                             index=EXPENSE_CATEGORIES.index(ext_data.get("suggested_category","Miscellaneous"))
                                             if ext_data.get("suggested_category") in EXPENSE_CATEGORIES else 0)
                    r_desc    = st.text_area("Description", value=ext_data.get("description",""), height=70)
                    r_total   = st.number_input("Total Bill Amount (RM)", value=float(ext_data.get("total_amount",0)),
                                                min_value=0.0, step=0.01, format="%.2f")
                    r_share   = st.number_input("ULU's Share (RM)", value=float(ext_data.get("total_amount",0)),
                                                min_value=0.0, step=0.01, format="%.2f",
                                                help="Adjust if this receipt was split across multiple properties")
                    if st.form_submit_button("✅ Confirm & Save", use_container_width=True):
                        conn = get_db()
                        conn.execute(
                            "INSERT INTO personal_expenses (year,month,vendor,bill_date,category,description,total_amount,ulu_share,file_name) VALUES (?,?,?,?,?,?,?,?,?)",
                            (st.session_state["ulu_year"], st.session_state["ulu_month"],
                             r_vendor, r_date, r_cat, r_desc, r_total, r_share,
                             st.session_state.get("ulu_file_name",""))
                        )
                        conn.commit(); conn.close()
                        del st.session_state["ulu_extracted"]
                        st.success("Receipt saved!")
                        st.rerun()
            else:
                st.markdown(
                    "<div style='text-align:center;padding:50px 0;color:#A89F91;'>"
                    "<div style='font-size:2.5rem'>📋</div>"
                    "Upload and extract a receipt — it will appear here for review."
                    "</div>", unsafe_allow_html=True
                )
            st.markdown('</div>', unsafe_allow_html=True)

# ══════════════════════════════════════════════
# TAB 3 — MONTHLY P&L
# ══════════════════════════════════════════════
with tab3:
    ym_list3 = get_year_month_list()
    ym_labels3 = [f"{MONTHS[m-1]} {y} ({operation_year(y,m)})" for y,m in ym_list3]
    sel_idx3 = st.selectbox("Select Month", range(len(ym_labels3)), format_func=lambda i: ym_labels3[i], key="pl_month")
    sel_year3, sel_month3 = ym_list3[sel_idx3]

    summary = get_monthly_summary(sel_year3, sel_month3)

    # Metrics row
    c1,c2,c3,c4 = st.columns(4)
    c1.markdown(f'<div class="metric-box"><div class="metric-label">Gross Income</div><div class="metric-value">RM {summary["gross_income"]:,.0f}</div></div>', unsafe_allow_html=True)
    c2.markdown(f'<div class="metric-box mid"><div class="metric-label">Gross Op Cost</div><div class="metric-value">RM {summary["gross_op_cost"]:,.0f}</div></div>', unsafe_allow_html=True)
    c3.markdown(f'<div class="metric-box mid"><div class="metric-label">Co-Host ({summary["cohost_pct"]:.0f}%)</div><div class="metric-value">RM {summary["cohost_share"]:,.0f}</div></div>', unsafe_allow_html=True)
    c4.markdown(f'<div class="metric-box accent"><div class="metric-label">Your Net Profit</div><div class="metric-value">RM {summary["owner_share"]:,.0f}</div></div>', unsafe_allow_html=True)

    st.markdown("<div style='height:10px'></div>", unsafe_allow_html=True)

    col_pl, col_bk = st.columns([1,1], gap="large")

    with col_pl:
        st.markdown('<div class="card">', unsafe_allow_html=True)
        st.markdown('<p class="card-title">P&L Statement</p>', unsafe_allow_html=True)
        pl_data = {
            "Description": [
                "Gross Income",
                "  Manager Expenses",
                "  Personal Expenses (ULU Share)",
                "Gross Operating Cost",
                "Net Profit Before Sharing",
                f"  Co-Host Share ({summary['cohost_pct']:.0f}%)",
                "Owner Net Profit"
            ],
            "Amount (RM)": [
                f"{summary['gross_income']:,.2f}",
                f"({summary['mgr_expenses']:,.2f})",
                f"({summary['personal_expenses']:,.2f})",
                f"({summary['gross_op_cost']:,.2f})",
                f"{summary['net_before_sharing']:,.2f}",
                f"({summary['cohost_share']:,.2f})",
                f"{summary['owner_share']:,.2f}",
            ]
        }
        st.dataframe(pd.DataFrame(pl_data), use_container_width=True, hide_index=True)

        # Nett profit %
        if summary["gross_income"] > 0:
            nett_pct = (summary["owner_share"] / summary["gross_income"]) * 100
            st.metric("Nett Profit %", f"{nett_pct:.1f}%")
        st.markdown('</div>', unsafe_allow_html=True)

    with col_bk:
        # Booking type breakdown
        conn = get_db()
        bks3 = conn.execute("SELECT * FROM bookings WHERE year=? AND month=? ORDER BY checkin", (sel_year3, sel_month3)).fetchall()
        conn.close()

        st.markdown('<div class="card">', unsafe_allow_html=True)
        st.markdown('<p class="card-title">Booking Mix</p>', unsafe_allow_html=True)
        if bks3:
            whole = [b for b in bks3 if b["room_type"] == "WHOLE"]
            mbed  = [b for b in bks3 if b["room_type"] == "MBED"]
            total_bk = len(bks3)
            w_pct = len(whole)/total_bk*100 if total_bk else 0
            m_pct = len(mbed)/total_bk*100 if total_bk else 0

            bm_data = {
                "Type": ["Whole Villa","Master Bed","TOTAL"],
                "Bookings": [len(whole), len(mbed), total_bk],
                "%": [f"{w_pct:.0f}%", f"{m_pct:.0f}%", "100%"],
                "Income (RM)": [
                    f"{sum(float(b['amount'] or 0) for b in whole):,.2f}",
                    f"{sum(float(b['amount'] or 0) for b in mbed):,.2f}",
                    f"{summary['gross_income']:,.2f}",
                ]
            }
            st.dataframe(pd.DataFrame(bm_data), use_container_width=True, hide_index=True)

            total_nights = sum(int(b["nights"] or 0) for b in bks3)
            airbnb_bk = len([b for b in bks3 if b["source"] == "AIRBNB"])
            direct_bk = len([b for b in bks3 if b["source"] == "DIRECT"])
            st.metric("Total Nights", total_nights)
            st.caption(f"Airbnb: {airbnb_bk} bookings · Direct: {direct_bk} bookings")
        else:
            st.info("No bookings this month.")
        st.markdown('</div>', unsafe_allow_html=True)

    # Download report
    st.markdown("---")
    if st.button("📄 Download Monthly PDF Report"):
        conn = get_db()
        bks_r   = conn.execute("SELECT * FROM bookings WHERE year=? AND month=? ORDER BY checkin", (sel_year3, sel_month3)).fetchall()
        mes_r   = conn.execute("SELECT * FROM manager_expenses WHERE year=? AND month=? ORDER BY id", (sel_year3, sel_month3)).fetchall()
        pes_r   = conn.execute("SELECT * FROM personal_expenses WHERE year=? AND month=? ORDER BY id", (sel_year3, sel_month3)).fetchall()
        conn.close()
        pdf = generate_monthly_report(sel_year3, sel_month3, summary, bks_r, mes_r, pes_r)
        fname = f"ULU_{MONTHS[sel_month3-1]}{sel_year3}_{operation_year(sel_year3,sel_month3)}.pdf"
        st.download_button("⬇️ Download PDF", data=pdf, file_name=fname, mime="application/pdf", use_container_width=True)

# ══════════════════════════════════════════════
# TAB 4 — YEARLY DASHBOARD
# ══════════════════════════════════════════════
with tab4:
    conn = get_db()
    bk_years = conn.execute("SELECT DISTINCT year FROM bookings").fetchall()
    ex_years = conn.execute("SELECT DISTINCT year FROM manager_expenses").fetchall()
    pe_years = conn.execute("SELECT DISTINCT year FROM personal_expenses").fetchall()
    conn.close()
    all_years = sorted(set(
        [r["year"] for r in bk_years] +
        [r["year"] for r in ex_years] +
        [r["year"] for r in pe_years]
    ))
    years_with_data = [{"year": y} for y in all_years]

    inv_cost = float(get_setting("investment_cost") or 410000)
    cohost_pct_setting = float(get_setting("cohost_pct") or 30)

    # ── ALL-TIME SUMMARY ───────────────────────────────────────────────────────
    st.markdown('<div class="card">', unsafe_allow_html=True)
    st.markdown('<p class="card-title">📊 All-Time Performance Summary</p>', unsafe_allow_html=True)

    # Compute all-time figures across all years
    conn = get_db()
    at_gross   = conn.execute("SELECT COALESCE(SUM(amount),0) as t FROM bookings").fetchone()["t"]
    at_mgr     = conn.execute("SELECT COALESCE(SUM(amount),0) as t FROM manager_expenses").fetchone()["t"]
    at_personal= conn.execute("SELECT COALESCE(SUM(ulu_share),0) as t FROM personal_expenses").fetchone()["t"]
    at_nights  = conn.execute("SELECT COALESCE(SUM(nights),0) as t FROM bookings").fetchone()["t"]
    at_bookings= conn.execute("SELECT COUNT(*) as t FROM bookings").fetchone()["t"]
    at_whole   = conn.execute("SELECT COUNT(*) as t FROM bookings WHERE room_type='WHOLE'").fetchone()["t"]
    at_mbed    = conn.execute("SELECT COUNT(*) as t FROM bookings WHERE room_type='MBED'").fetchone()["t"]
    # Months in operation
    start_y = int(get_setting("operation_start_year") or 2024)
    start_m = int(get_setting("operation_start_month") or 7)
    now = datetime.datetime.now()
    months_ops = (now.year - start_y) * 12 + (now.month - start_m) + 1
    conn.close()

    at_opex        = at_mgr + at_personal
    at_net_before  = at_gross - at_opex
    at_cohost      = max(0, at_net_before * cohost_pct_setting / 100)
    at_owner       = max(0, at_net_before - at_cohost)
    at_roi         = (at_owner / inv_cost * 100) if inv_cost else 0
    at_avg_night   = at_gross / at_nights if at_nights else 0
    at_avg_monthly = at_owner / months_ops if months_ops else 0
    payback_months = (inv_cost / at_avg_monthly) if at_avg_monthly > 0 else 0

    # Row 1 — Income metrics
    r1c1, r1c2, r1c3, r1c4 = st.columns(4)
    r1c1.markdown(f'<div class="metric-box"><div class="metric-label">Total Gross Income</div><div class="metric-value">RM {at_gross:,.0f}</div></div>', unsafe_allow_html=True)
    r1c2.markdown(f'<div class="metric-box mid"><div class="metric-label">Total OpEx Paid</div><div class="metric-value">RM {at_opex:,.0f}</div></div>', unsafe_allow_html=True)
    r1c3.markdown(f'<div class="metric-box mid"><div class="metric-label">Co-Host Paid (Azary)</div><div class="metric-value">RM {at_cohost:,.0f}</div></div>', unsafe_allow_html=True)
    r1c4.markdown(f'<div class="metric-box accent"><div class="metric-label">Your Net Profit</div><div class="metric-value">RM {at_owner:,.0f}</div></div>', unsafe_allow_html=True)

    st.markdown("<div style='height:10px'></div>", unsafe_allow_html=True)

    # Row 2 — ROI metrics
    r2c1, r2c2, r2c3, r2c4 = st.columns(4)
    r2c1.markdown(f'<div class="metric-box mid"><div class="metric-label">Investment Cost</div><div class="metric-value">RM {inv_cost:,.0f}</div></div>', unsafe_allow_html=True)
    r2c2.markdown(f'<div class="metric-box accent"><div class="metric-label">Cumulative ROI</div><div class="metric-value">{at_roi:.1f}%</div></div>', unsafe_allow_html=True)
    r2c3.markdown(f'<div class="metric-box mid"><div class="metric-label">Months Operating</div><div class="metric-value">{months_ops}</div></div>', unsafe_allow_html=True)
    r2c4.markdown(f'<div class="metric-box"><div class="metric-label">Avg Net / Month</div><div class="metric-value">RM {at_avg_monthly:,.0f}</div></div>', unsafe_allow_html=True)

    st.markdown("<div style='height:10px'></div>", unsafe_allow_html=True)

    # Row 3 — Operational metrics
    r3c1, r3c2, r3c3, r3c4 = st.columns(4)
    r3c1.markdown(f'<div class="metric-box mid"><div class="metric-label">Total Bookings</div><div class="metric-value">{at_bookings}</div></div>', unsafe_allow_html=True)
    r3c2.markdown(f'<div class="metric-box mid"><div class="metric-label">Total Nights Sold</div><div class="metric-value">{int(at_nights)}</div></div>', unsafe_allow_html=True)
    r3c3.markdown(f'<div class="metric-box mid"><div class="metric-label">Avg Rate / Night</div><div class="metric-value">RM {at_avg_night:,.0f}</div></div>', unsafe_allow_html=True)
    r3c4.markdown(f'<div class="metric-box accent"><div class="metric-label">Est. Payback Period</div><div class="metric-value">{payback_months:.0f} mths</div></div>', unsafe_allow_html=True)

    st.markdown("<div style='height:10px'></div>", unsafe_allow_html=True)

    # Booking mix all-time
    if at_bookings > 0:
        st.markdown(f"""
        <div style='display:flex;gap:16px;flex-wrap:wrap;'>
            <span><b>Whole Villa:</b> {at_whole} bookings ({at_whole/at_bookings*100:.0f}%)</span>
            <span>·</span>
            <span><b>Master Bed:</b> {at_mbed} bookings ({at_mbed/at_bookings*100:.0f}%)</span>
            <span>·</span>
            <span><b>Co-Host Rate:</b> {cohost_pct_setting:.0f}%</span>
        </div>
        """, unsafe_allow_html=True)

    st.markdown('</div>', unsafe_allow_html=True)
    st.markdown("<div style='height:16px'></div>", unsafe_allow_html=True)

    # ── YEAR SELECTOR (existing) ───────────────────────────────────────────────
    available_years = [r["year"] for r in years_with_data] if years_with_data else [datetime.datetime.now().year]
    sel_year4 = st.selectbox("Select Year for Detailed View", available_years, index=len(available_years)-1, key="yr_select")

    monthly_data = get_yearly_summary(sel_year4)

    # Build yearly totals
    yearly_income  = sum(v["gross_income"] for v in monthly_data.values())
    yearly_op_cost = sum(v["gross_op_cost"] for v in monthly_data.values())
    yearly_net     = sum(v["owner_share"] for v in monthly_data.values())
    yearly_roi     = (yearly_net / inv_cost * 100) if inv_cost else 0

    # Year metrics
    c1,c2,c3,c4 = st.columns(4)
    c1.markdown(f'<div class="metric-box"><div class="metric-label">Yearly Gross Income</div><div class="metric-value">RM {yearly_income:,.0f}</div></div>', unsafe_allow_html=True)
    c2.markdown(f'<div class="metric-box mid"><div class="metric-label">Yearly Op Cost</div><div class="metric-value">RM {yearly_op_cost:,.0f}</div></div>', unsafe_allow_html=True)
    c3.markdown(f'<div class="metric-box accent"><div class="metric-label">Yearly Net Profit</div><div class="metric-value">RM {yearly_net:,.0f}</div></div>', unsafe_allow_html=True)
    c4.markdown(f'<div class="metric-box mid"><div class="metric-label">Yearly ROI</div><div class="metric-value">{yearly_roi:.1f}%</div></div>', unsafe_allow_html=True)

    st.markdown("<div style='height:12px'></div>", unsafe_allow_html=True)

    # Monthly breakdown table
    st.markdown('<div class="card">', unsafe_allow_html=True)
    st.markdown('<p class="card-title">Monthly Breakdown</p>', unsafe_allow_html=True)

    rows = []
    for month_num in sorted(monthly_data.keys()):
        v = monthly_data[month_num]
        nett_pct = (v["owner_share"] / v["gross_income"] * 100) if v["gross_income"] > 0 else 0
        rows.append({
            "Month": f"{MONTHS[month_num-1]} {sel_year4}",
            "Op Year": operation_year(sel_year4, month_num),
            "Gross Income": f"{v['gross_income']:,.2f}",
            "Op Cost": f"{v['gross_op_cost']:,.2f}",
            "Net Before Share": f"{v['net_before_sharing']:,.2f}",
            "Co-Host": f"{v['cohost_share']:,.2f}",
            "Owner Net": f"{v['owner_share']:,.2f}",
            "Net %": f"{nett_pct:.1f}%",
        })

    if rows:
        st.dataframe(pd.DataFrame(rows), use_container_width=True, hide_index=True)
    else:
        st.info("No data for this year yet.")
    st.markdown('</div>', unsafe_allow_html=True)

    # Booking mix for year
    st.markdown('<div class="card">', unsafe_allow_html=True)
    st.markdown('<p class="card-title">Yearly Booking Analysis</p>', unsafe_allow_html=True)
    conn = get_db()
    all_bks = conn.execute("SELECT * FROM bookings WHERE year=?", (sel_year4,)).fetchall()
    conn.close()

    if all_bks:
        total_bks    = len(all_bks)
        whole_bks    = [b for b in all_bks if b["room_type"] == "WHOLE"]
        mbed_bks     = [b for b in all_bks if b["room_type"] == "MBED"]
        airbnb_bks   = len([b for b in all_bks if b["source"] == "AIRBNB"])
        direct_bks   = len([b for b in all_bks if b["source"] == "DIRECT"])
        total_nights = sum(int(b["nights"] or 0) for b in all_bks)
        whole_nights = sum(int(b["nights"] or 0) for b in whole_bks)
        mbed_nights  = sum(int(b["nights"] or 0) for b in mbed_bks)
        whole_income = sum(float(b["amount"] or 0) for b in whole_bks)
        mbed_income  = sum(float(b["amount"] or 0) for b in mbed_bks)

        ca, cb, cc, cd, ce = st.columns(5)
        ca.metric("Total Bookings", total_bks)
        cb.metric("Whole Villa", f"{len(whole_bks)} ({len(whole_bks)/total_bks*100:.0f}%)")
        cc.metric("Master Bed", f"{len(mbed_bks)} ({len(mbed_bks)/total_bks*100:.0f}%)")
        cd.metric("Airbnb", f"{airbnb_bks} ({airbnb_bks/total_bks*100:.0f}%)")
        ce.metric("Total Nights", total_nights)

        avg_per_night = yearly_income / total_nights if total_nights else 0
        st.markdown(f"<span style='color:#1C1C1A;font-size:1rem'>Average rate per night: <b>{fmt_myr(avg_per_night)}</b></span>", unsafe_allow_html=True)

        st.markdown("---")
        st.markdown("**Room Type Breakdown — Bookings, Nights & Revenue**")
        mix_data = {
            "Type": ["Whole Villa", "Master Bed (M-Bed)", "TOTAL"],
            "Bookings": [len(whole_bks), len(mbed_bks), total_bks],
            "% Bookings": [
                f"{len(whole_bks)/total_bks*100:.0f}%" if total_bks else "—",
                f"{len(mbed_bks)/total_bks*100:.0f}%" if total_bks else "—",
                "100%"
            ],
            "Nights": [whole_nights, mbed_nights, total_nights],
            "% Nights": [
                f"{whole_nights/total_nights*100:.0f}%" if total_nights else "—",
                f"{mbed_nights/total_nights*100:.0f}%" if total_nights else "—",
                "100%"
            ],
            "Revenue (RM)": [
                f"{whole_income:,.2f}", f"{mbed_income:,.2f}", f"{yearly_income:,.2f}"
            ],
            "Avg/Night (RM)": [
                f"{whole_income/whole_nights:,.2f}" if whole_nights else "—",
                f"{mbed_income/mbed_nights:,.2f}" if mbed_nights else "—",
                f"{avg_per_night:,.2f}"
            ],
        }
        st.dataframe(pd.DataFrame(mix_data), use_container_width=True, hide_index=True)
    else:
        st.info("No booking data for this year.")
    st.markdown('</div>', unsafe_allow_html=True)

    # Cumulative ROI since inception
    st.markdown('<div class="card">', unsafe_allow_html=True)
    st.markdown('<p class="card-title">ROI Tracker vs Investment Cost</p>', unsafe_allow_html=True)
    conn = get_db()
    _bk_yrs = conn.execute("SELECT DISTINCT year FROM bookings").fetchall()
    _ex_yrs = conn.execute("SELECT DISTINCT year FROM manager_expenses").fetchall()
    conn.close()
    all_years = sorted(set(
        [r["year"] for r in _bk_yrs] + [r["year"] for r in _ex_yrs]
    ))
    all_years = [{"year": y} for y in all_years]

    cum_profit = 0
    roi_rows = []
    for yr_row in all_years:
        yr = yr_row["year"]
        yd = get_yearly_summary(yr)
        yr_net = sum(v["owner_share"] for v in yd.values())
        yr_inc = sum(v["gross_income"] for v in yd.values())
        cum_profit += yr_net
        cum_roi = (cum_profit / inv_cost * 100) if inv_cost else 0
        roi_rows.append({
            "Year": yr,
            "Op Year": operation_year(yr, 7),
            "Yearly Income": fmt_myr(yr_inc),
            "Yearly Net Profit": fmt_myr(yr_net),
            "Cumulative Profit": fmt_myr(cum_profit),
            "Cumulative ROI": f"{cum_roi:.2f}%",
        })
    if roi_rows:
        st.dataframe(pd.DataFrame(roi_rows), use_container_width=True, hide_index=True)
        st.markdown(f"<span style='color:#1C1C1A;font-size:1rem'>Investment cost: <b>{fmt_myr(inv_cost)}</b></span>", unsafe_allow_html=True)
    st.markdown('</div>', unsafe_allow_html=True)

# ══════════════════════════════════════════════
# TAB 5 — ALL RECORDS
# ══════════════════════════════════════════════
with tab5:
    st.markdown("Browse and audit all records across all months.")
    record_type = st.radio("View", ["Bookings","Manager Expenses","Personal Receipts"], horizontal=True)

    conn = get_db()
    if record_type == "Bookings":
        rows = conn.execute("SELECT * FROM bookings ORDER BY year DESC, month DESC, checkin").fetchall()
        if rows:
            df = pd.DataFrame([dict(r) for r in rows])
            df = df[["id","year","month","guest_name","room_type","checkin","checkout","nights","source","amount","notes"]]
            df.columns = ["ID","Year","Month","Guest","Type","Check-in","Check-out","Nights","Source","Amount (RM)","Notes"]
            df["Amount (RM)"] = df["Amount (RM)"].apply(lambda x: f"{float(x):,.2f}")
            df["Month"] = df["Month"].apply(lambda m: MONTHS[int(m)-1])
            st.dataframe(df, use_container_width=True, hide_index=True)
            st.metric("Total Income All Time", fmt_myr(sum(float(r["amount"] or 0) for r in rows)))
        else:
            st.info("No bookings yet.")

    elif record_type == "Manager Expenses":
        rows = conn.execute("SELECT * FROM manager_expenses ORDER BY year DESC, month DESC, id").fetchall()
        if rows:
            df = pd.DataFrame([dict(r) for r in rows])
            df = df[["id","year","month","expense_item","vendor","amount","notes"]]
            df.columns = ["ID","Year","Month","Item","Vendor","Amount (RM)","Notes"]
            df["Amount (RM)"] = df["Amount (RM)"].apply(lambda x: f"{float(x):,.2f}")
            df["Month"] = df["Month"].apply(lambda m: MONTHS[int(m)-1])
            st.dataframe(df, use_container_width=True, hide_index=True)
            st.metric("Total Manager Expenses All Time", fmt_myr(sum(float(r["amount"] or 0) for r in rows)))
        else:
            st.info("No manager expenses yet.")

    else:
        rows = conn.execute("SELECT * FROM personal_expenses ORDER BY year DESC, month DESC, id").fetchall()
        if rows:
            df = pd.DataFrame([dict(r) for r in rows])
            df = df[["id","year","month","vendor","bill_date","category","description","total_amount","ulu_share","file_name"]]
            df.columns = ["ID","Year","Month","Vendor","Date","Category","Description","Total Bill","ULU Share","Source"]
            df["Total Bill"] = df["Total Bill"].apply(lambda x: f"{float(x):,.2f}")
            df["ULU Share"]  = df["ULU Share"].apply(lambda x: f"{float(x):,.2f}")
            df["Month"] = df["Month"].apply(lambda m: MONTHS[int(m)-1])
            st.dataframe(df, use_container_width=True, hide_index=True)
            st.metric("Total Personal Expenses (ULU Share) All Time",
                      fmt_myr(sum(float(r["ulu_share"] or 0) for r in rows)))
        else:
            st.info("No personal receipts yet.")
    conn.close()


# ══════════════════════════════════════════════
# TAB 6 — CAPEX TRACKER
# ══════════════════════════════════════════════
with tab6:
    st.markdown('<div class="card">', unsafe_allow_html=True)
    st.markdown('<p class="card-title">🏗️ Capital Expenditure (CapEx) Tracker — ULU 1 Upgrade</p>', unsafe_allow_html=True)
    st.caption("Track all capital improvement costs for ULU Mahsuri Villa. "
               "CapEx = purchases that improve or add long-term value (furniture, renovation, equipment above RM1,000).")

    capex_add, capex_view = st.tabs(["➕ Add CapEx Item", "📋 All CapEx Items"])

    with capex_add:
        st.markdown("#### Add New CapEx Item")
        st.caption("For AI extraction, upload your receipt or invoice image/PDF.")

        # ── AI Extraction ─────────────────────────────────────────────────────
        with st.expander("📎 Upload Receipt/Invoice — AI will extract details", expanded=True):
            capex_upload = st.file_uploader("Drop invoice or receipt",
                type=["pdf","png","jpg","jpeg"], key="capex_upload")
            if capex_upload:
                cur_name = st.session_state.get("capex_upload_name","")
                if capex_upload.name != cur_name:
                    st.session_state["capex_bytes"] = capex_upload.read()
                    st.session_state["capex_name"]  = capex_upload.name
                    st.session_state["capex_upload_name"] = capex_upload.name
                    st.session_state["capex_saved"] = False

            if st.session_state.get("capex_bytes"):
                st.success(f"✓ File ready: {st.session_state.get('capex_name','')} "
                           f"({len(st.session_state['capex_bytes']):,} bytes)")
                if not st.session_state.get("ulu_api_key"):
                    st.warning("Enter your Claude API key in the sidebar first.")
                elif st.button("🤖 Extract with AI", key="ai_capex", type="primary"):
                    with st.spinner("AI reading document..."):
                        try:
                            import anthropic as _ant, base64 as _b64, json as _json
                            fbytes = st.session_state["capex_bytes"]
                            fname  = st.session_state["capex_name"].lower()
                            b64    = _b64.standard_b64encode(fbytes).decode()
                            if fname.endswith(".pdf"):
                                mtype, stype = "application/pdf", "document"
                            else:
                                ext = fname.split(".")[-1]
                                mtype = f"image/{'jpeg' if ext in ['jpg','jpeg'] else ext}"
                                stype = "image"
                            cli = _ant.Anthropic(
                                api_key=st.session_state.get("ulu_api_key",""),
                                timeout=60.0
                            )
                            prompt = """Extract purchase details from this receipt or invoice.
Return ONLY a JSON object:
{"vendor": "supplier name", "description": "what was purchased", "amount": 1234.56,
 "purchase_date": "YYYY-MM-DD", "category": "best matching category"}
Category must be one of: Furniture & Furnishings (FF&E), Soft Furnishings & Decor,
Kitchen Equipment & Appliances, Bathroom Fittings & Fixtures, Electrical & Lighting,
Renovation & Structural Works, Outdoor & Landscaping, Technology & Smart Home,
Pool & Recreation Equipment, Miscellaneous CapEx.
Return ONLY the JSON."""
                            src = {"type":"base64","media_type":mtype,"data":b64}
                            msg = [{"type":stype,"source":src},{"type":"text","text":prompt}] if stype=="image" else [{"type":"document","source":src},{"type":"text","text":prompt}]
                            resp = cli.messages.create(model="claude-sonnet-4-6", max_tokens=400,
                                messages=[{"role":"user","content":msg}])
                            raw = resp.content[0].text.strip()
                            if raw.startswith("```"): raw = raw.split("```")[1]; raw = raw[4:] if raw.startswith("json") else raw
                            ext_data = _json.loads(raw.strip())
                            st.session_state["capex_vendor"]  = ext_data.get("vendor","")
                            st.session_state["capex_desc"]    = ext_data.get("description","")
                            st.session_state["capex_amount"]  = float(ext_data.get("amount",0))
                            st.session_state["capex_date"]    = ext_data.get("purchase_date","")
                            st.session_state["capex_cat"]     = ext_data.get("category","Miscellaneous CapEx")
                            # Clear conflicting widget keys so values reload from session state
                            for wk in ["cx_date","cx_vendor","cx_desc","cx_cat","cx_amount"]:
                                st.session_state.pop(wk, None)
                            st.success(f"✓ Extracted — RM {ext_data.get('amount',0):,.2f}")
                            st.rerun()
                        except Exception as e:
                            st.error(f"Extraction failed: {e}")

        # ── Form ─────────────────────────────────────────────────────────────
        c1, c2 = st.columns(2)
        cx_date   = c1.text_input("Purchase Date (YYYY-MM-DD)",
            value=st.session_state.get("capex_date",""))
        cx_vendor = c2.text_input("Vendor / Supplier",
            value=st.session_state.get("capex_vendor",""))
        cx_desc   = st.text_area("Description of Works / Item",
            value=st.session_state.get("capex_desc",""), height=70)
        c3, c4 = st.columns(2)

        cat_default = st.session_state.get("capex_cat","Furniture & Furnishings (FF&E)")
        cat_idx = CAPEX_CATEGORIES.index(cat_default) if cat_default in CAPEX_CATEGORIES else 0
        cx_cat    = c3.selectbox("CapEx Category", CAPEX_CATEGORIES, index=cat_idx)
        cx_amount = c4.number_input("Amount (MYR)", min_value=0.0,
            value=st.session_state.get("capex_amount",0.0),
            step=100.0, format="%.2f")
        cx_life   = c3.number_input("Useful Life (years)", min_value=1, max_value=50, value=5)
        cx_notes  = c4.text_input("Notes")

        if st.button("💾 Save CapEx Item", type="primary", key="save_capex"):
            if not cx_desc.strip() or cx_amount <= 0:
                st.error("Description and amount are required.")
            else:
                receipt_path   = ""
                file_name_save = ""

                if st.session_state.get("capex_bytes"):
                    vendor_clean   = (cx_vendor or "Unknown").replace("/","_")[:20]
                    file_name_save = f"{cx_date}_{vendor_clean}_{st.session_state.get('capex_name','receipt')}"

                    # Detect if running locally (not on Streamlit Cloud)
                    is_cloud = os.environ.get("STREAMLIT_SHARING_MODE") or \
                               os.environ.get("HOME","").startswith("/home/appuser") or \
                               "/mount/src" in os.path.abspath(__file__)

                    if not is_cloud:
                        # Local — save to CapEx Receipts folder
                        scan_folder = os.path.join(os.path.dirname(os.path.abspath(__file__)),
                                                   "CapEx Receipts")
                        os.makedirs(scan_folder, exist_ok=True)
                        receipt_path = os.path.join(scan_folder, file_name_save)
                        with open(receipt_path, "wb") as _f:
                            _f.write(st.session_state["capex_bytes"])
                    else:
                        # Cloud — store filename only, no local disk
                        receipt_path = file_name_save

                conn = get_db()
                db_error = None
                try:
                    conn.execute("""INSERT INTO capex_items
                        (purchase_date,vendor,description,category,amount,useful_life_years,file_name,receipt_path,notes)
                        VALUES (?,?,?,?,?,?,?,?,?)""",
                        (cx_date, cx_vendor, cx_desc.strip(), cx_cat,
                         cx_amount, cx_life, file_name_save, receipt_path, cx_notes))
                    conn.commit()
                except Exception as e:
                    db_error = str(e)
                finally:
                    conn.close()

                if db_error:
                    st.error(f"❌ Save failed — database error: {db_error}")
                    st.info("Tip: Check Supabase dashboard → capex_items table to confirm columns exist.")
                else:
                    for k in ["capex_bytes","capex_name","capex_vendor","capex_desc",
                              "capex_amount","capex_date","capex_cat","capex_upload_name"]:
                        st.session_state.pop(k, None)
                    st.success("✓ CapEx item saved.")
                    st.rerun()

    with capex_view:
        conn = get_db()
        capex_rows = conn.execute(
            "SELECT * FROM capex_items ORDER BY purchase_date DESC"
        ).fetchall()
        conn.close()

        if capex_rows:
            total_capex = sum(float(r["amount"]) for r in capex_rows)

            # Summary by category
            cat_totals = {}
            for r in capex_rows:
                cat_totals[r["category"]] = cat_totals.get(r["category"],0) + float(r["amount"])

            st.metric("Total CapEx — ULU 1 Upgrade", fmt_myr(total_capex))

            # Category breakdown
            cat_rows = [{"Category": k, "Amount (MYR)": fmt_myr(v), "% of Total": f"{v/total_capex*100:.1f}%"}
                        for k,v in sorted(cat_totals.items(), key=lambda x:-x[1])]
            st.dataframe(pd.DataFrame(cat_rows), use_container_width=True, hide_index=True)
            st.divider()

            # Full list with delete
            st.subheader("All CapEx Items")
            for r in [dict(r) for r in capex_rows]:
                rc1,rc2,rc3,rc4,rc5,rc6 = st.columns([2,3,2,2,1,1])
                rc1.markdown(f"<span style='color:#1C1C1A;font-size:0.9rem'>{r.get('purchase_date','')}</span>", unsafe_allow_html=True)
                rc2.markdown(f"<span style='color:#1C1C1A;font-size:0.9rem'>{(r.get('description','') or '')[:40]}</span>", unsafe_allow_html=True)
                rc3.markdown(f"<span style='color:#1C1C1A;font-size:0.9rem'>{(r.get('category','') or '')[:25]}</span>", unsafe_allow_html=True)
                rc4.markdown(f"<span style='color:#1C1C1A;font-size:0.9rem;font-weight:600'>{fmt_myr(r.get('amount',0))}</span>", unsafe_allow_html=True)
                # Receipt file — use file_name or receipt_path
                sp = r.get("receipt_path","") or r.get("scan_path","") or r.get("file_name","")
                if sp and os.path.exists(sp):
                    with open(sp,"rb") as _f:
                        ext = sp.split(".")[-1].lower()
                        mime = "application/pdf" if ext=="pdf" else f"image/{ext}"
                        rc5.download_button("📎", data=_f.read(),
                            file_name=os.path.basename(sp), mime=mime,
                            key=f"cx_dl_{r['id']}")
                elif sp:
                    rc5.caption("📎")  # file recorded but not on this machine
                else:
                    rc5.caption("—")
                if rc6.button("🗑", key=f"cx_del_{r['id']}"):
                    conn = get_db()
                    conn.execute("DELETE FROM capex_items WHERE id=?", (r["id"],))
                    conn.commit(); conn.close()
                    st.rerun()
        else:
            st.info("No CapEx items recorded yet. Add items using the Add tab.")
    st.markdown('</div>', unsafe_allow_html=True)


# ══════════════════════════════════════════════
# TAB 7 — OPEX BREAKDOWN
# ══════════════════════════════════════════════
with tab7:
    st.markdown('<div class="card">', unsafe_allow_html=True)
    st.markdown('<p class="card-title">📋 OpEx Breakdown — All Operating Expenses</p>',
                unsafe_allow_html=True)
    st.caption("Full view of all operating expenses: Manager submissions + Owner personal receipts (ULU share).")

    opex_tab_a, opex_tab_b = st.tabs(["🏷️ Classify Manager Expenses", "📊 OpEx Summary"])

    with opex_tab_a:
        # Filter selector
        fy1, fy2 = st.columns(2)
        opex_year  = fy1.selectbox("Year",  list(range(datetime.datetime.now().year, 2023, -1)), key="opex_yr")
        opex_month = fy2.selectbox("Month", ["All"] + MONTHS, key="opex_mo")

        conn = get_db()
        if opex_month == "All":
            exp_rows = conn.execute(
                "SELECT * FROM manager_expenses WHERE year=? ORDER BY month, id",
                (opex_year,)
            ).fetchall()
        else:
            mo_num = MONTHS.index(opex_month) + 1
            exp_rows = conn.execute(
                "SELECT * FROM manager_expenses WHERE year=? AND month=? ORDER BY id",
                (opex_year, mo_num)
            ).fetchall()
        conn.close()

        if not exp_rows:
            st.info("No manager expenses found for this period.")
        else:
            uncat = [r for r in exp_rows if not r["opex_category"] or r["opex_category"] == "Uncategorised"]
            if uncat:
                st.warning(f"⚠️ {len(uncat)} expense(s) still uncategorised. Please classify them below.")
            else:
                st.success(f"✓ All {len(exp_rows)} manager expenses are classified.")

            st.divider()
            for r in [dict(r) for r in exp_rows]:
                cur_cat = r.get("opex_category") or "Uncategorised"
                cat_idx = OPEX_CATEGORIES.index(cur_cat) if cur_cat in OPEX_CATEGORIES else 0
                highlight = "background:#fff8e1;padding:8px;border-radius:6px;margin-bottom:6px" if cur_cat == "Uncategorised" else "padding:4px;margin-bottom:4px"
                st.markdown(f"<div style='{highlight}'>", unsafe_allow_html=True)
                oc1,oc2,oc3,oc4 = st.columns([2,3,3,1])
                oc1.markdown(f"<span style='color:#1C1C1A;font-size:1rem'>{MONTHS[r['month']-1]} {r['year']}</span>", unsafe_allow_html=True)
                oc2.markdown(f"<span style='color:#1C1C1A;font-size:1rem'><b>{r['expense_item']}</b> — {r['vendor'] or ''}</span>", unsafe_allow_html=True)
                oc3.markdown(f"<span style='color:#1C1C1A;font-size:1rem;font-weight:600'>{fmt_myr(r['amount'])}</span>", unsafe_allow_html=True)
                new_cat = oc4.selectbox("Category", OPEX_CATEGORIES,
                    index=cat_idx, key=f"opex_cat_{r['id']}",
                    label_visibility="collapsed")
                st.markdown("</div>", unsafe_allow_html=True)

                if new_cat != cur_cat:
                    conn = get_db()
                    conn.execute("UPDATE manager_expenses SET opex_category=? WHERE id=?",
                                 (new_cat, r["id"]))
                    conn.commit(); conn.close()

        # Personal expenses for same period — with edit and delete
        st.divider()
        st.markdown("**Owner Personal Receipts (ULU Share) — same period:**")
        conn = get_db()
        if opex_month == "All":
            pe_rows = conn.execute(
                "SELECT * FROM personal_expenses WHERE year=? ORDER BY month, id",
                (opex_year,)
            ).fetchall()
        else:
            mo_num = MONTHS.index(opex_month) + 1
            pe_rows = conn.execute(
                "SELECT * FROM personal_expenses WHERE year=? AND month=? ORDER BY id",
                (opex_year, mo_num)
            ).fetchall()
        conn.close()

        if not pe_rows:
            st.caption("No personal receipts for this period.")
        else:
            pe_total = sum(float(r["ulu_share"] or 0) for r in pe_rows)
            for r in [dict(r) for r in pe_rows]:
                pe_highlight = "padding:4px;margin-bottom:4px"
                st.markdown(f"<div style='{pe_highlight}'>", unsafe_allow_html=True)
                pc1, pc2, pc3, pc4, pc5, pc6 = st.columns([1.5, 2.5, 2, 2, 1.5, 0.5])
                pc1.markdown(f"<span style='color:#1C1C1A;font-size:1rem'>{MONTHS[r['month']-1]} {r['year']}</span>", unsafe_allow_html=True)
                pc2.markdown(f"<span style='color:#1C1C1A;font-size:1rem'><b>{r['vendor']}</b></span>", unsafe_allow_html=True)
                pc3.markdown(f"<span style='color:#1C1C1A;font-size:0.9rem'>{r['description'][:35] if r['description'] else ''}</span>", unsafe_allow_html=True)
                pc4.markdown(f"<span style='color:#1C1C1A;font-size:1rem;font-weight:600'>{fmt_myr(r['ulu_share'])}</span>", unsafe_allow_html=True)
                # Category edit
                cur_pe_cat = r.get("category") or "Miscellaneous"
                pe_cat_idx = EXPENSE_CATEGORIES.index(cur_pe_cat) if cur_pe_cat in EXPENSE_CATEGORIES else 0
                new_pe_cat = pc5.selectbox(
                    "Cat", EXPENSE_CATEGORIES,
                    index=pe_cat_idx,
                    key=f"pe_cat_{r['id']}",
                    label_visibility="collapsed"
                )
                if new_pe_cat != cur_pe_cat:
                    conn = get_db()
                    conn.execute("UPDATE personal_expenses SET category=? WHERE id=?",
                                 (new_pe_cat, r["id"]))
                    conn.commit(); conn.close()
                # Delete button
                if pc6.button("🗑", key=f"pe_del_{r['id']}"):
                    conn = get_db()
                    conn.execute("DELETE FROM personal_expenses WHERE id=?", (r["id"],))
                    conn.commit(); conn.close()
                    st.rerun()
                st.markdown("</div>", unsafe_allow_html=True)
            st.markdown(f"**Personal Receipts Total (ULU Share): {fmt_myr(pe_total)}**")

    with opex_tab_b:
        sby1, sby2 = st.columns(2)
        sum_year  = sby1.selectbox("Year", list(range(datetime.datetime.now().year, 2023, -1)), key="sum_yr")
        sum_period = sby2.selectbox("Period", ["Full Year","Q1 (Jan-Mar)","Q2 (Apr-Jun)",
                                               "Q3 (Jul-Sep)","Q4 (Oct-Dec)"], key="sum_period")
        period_months = {
            "Full Year": list(range(1,13)),
            "Q1 (Jan-Mar)": [1,2,3], "Q2 (Apr-Jun)": [4,5,6],
            "Q3 (Jul-Sep)": [7,8,9], "Q4 (Oct-Dec)": [10,11,12]
        }[sum_period]

        conn = get_db()
        placeholders = ",".join("?" * len(period_months))
        all_mgr_exp = conn.execute(
            f"SELECT * FROM manager_expenses WHERE year=? AND month IN ({placeholders})",
            [sum_year] + period_months
        ).fetchall()
        all_pe_exp = conn.execute(
            f"SELECT * FROM personal_expenses WHERE year=? AND month IN ({placeholders})",
            [sum_year] + period_months
        ).fetchall()
        conn.close()

        if not all_mgr_exp and not all_pe_exp:
            st.info("No expense data for this period.")
        else:
            # Manager expenses by opex_category
            mgr_cat_totals = {}
            for r in all_mgr_exp:
                cat = r["opex_category"] or "Uncategorised"
                mgr_cat_totals[cat] = mgr_cat_totals.get(cat, 0) + float(r["amount"] or 0)

            # Personal expenses by category
            pe_cat_totals = {}
            for r in all_pe_exp:
                cat = r["category"] or "Miscellaneous"
                pe_cat_totals[cat] = pe_cat_totals.get(cat, 0) + float(r["ulu_share"] or 0)

            mgr_total = sum(mgr_cat_totals.values())
            pe_total  = sum(pe_cat_totals.values())
            total_opex = mgr_total + pe_total

            # Top metrics
            m1, m2, m3 = st.columns(3)
            m1.metric("Manager Expenses", fmt_myr(mgr_total))
            m2.metric("Personal Receipts (ULU Share)", fmt_myr(pe_total))
            m3.metric("Total OpEx", fmt_myr(total_opex))

            # Utilities vs Other
            utility_cats = [c for c in mgr_cat_totals if "Utilities" in c]
            utility_total = sum(mgr_cat_totals[c] for c in utility_cats)
            other_total   = total_opex - utility_total
            u1, u2 = st.columns(2)
            u1.metric("Utilities Total", fmt_myr(utility_total))
            u2.metric("Other OpEx", fmt_myr(other_total))
            st.divider()

            # Manager breakdown
            st.markdown("**Manager Expenses by Category:**")
            if mgr_cat_totals:
                mgr_rows = [{"OpEx Category": k,
                             "Amount (MYR)": fmt_myr(v),
                             "% of Total OpEx": f"{v/total_opex*100:.1f}%"}
                            for k,v in sorted(mgr_cat_totals.items(), key=lambda x:-x[1])]
                st.dataframe(pd.DataFrame(mgr_rows), use_container_width=True, hide_index=True)
            else:
                st.caption("No manager expenses.")

            # Personal breakdown
            st.markdown("**Owner Personal Receipts by Category:**")
            if pe_cat_totals:
                pe_rows_disp = [{"Category": k,
                                 "ULU Share (MYR)": fmt_myr(v),
                                 "% of Total OpEx": f"{v/total_opex*100:.1f}%"}
                                for k,v in sorted(pe_cat_totals.items(), key=lambda x:-x[1])]
                st.dataframe(pd.DataFrame(pe_rows_disp), use_container_width=True, hide_index=True)
            else:
                st.caption("No personal receipts.")

            st.info("💡 For LHDN purposes: Utilities, Housekeeping, Pool Maintenance, "
                    "Operation Items are fully deductible. "
                    "Renovation/structural works above RM1,000 should be classified as CapEx.")
    st.markdown('</div>', unsafe_allow_html=True)


# ══════════════════════════════════════════════
# TAB 8 — ACCOUNTANT
# ══════════════════════════════════════════════
with tab8:
    st.markdown('<div class="card">', unsafe_allow_html=True)
    st.markdown('<p class="card-title">📦 Accountant & Tax Agent Access</p>', unsafe_allow_html=True)

    acct_tab1, acct_tab2 = st.tabs(["📊 Export Reports", "🔐 Share Access"])

    with acct_tab1:
        st.subheader("Generate Financial Reports")
        er1, er2 = st.columns(2)
        exp_year   = er1.selectbox("Year", list(range(datetime.datetime.now().year, 2023, -1)), key="exp_yr")
        exp_period = er2.selectbox("Period", ["Full Year","Q1 (Jan-Mar)","Q2 (Apr-Jun)",
                                              "Q3 (Jul-Sep)","Q4 (Oct-Dec)"], key="exp_period")
        period_mos = {
            "Full Year": list(range(1,13)),
            "Q1 (Jan-Mar)": [1,2,3], "Q2 (Apr-Jun)": [4,5,6],
            "Q3 (Jul-Sep)": [7,8,9], "Q4 (Oct-Dec)": [10,11,12]
        }[exp_period]

        # Preview metrics
        conn = get_db()
        ph = ",".join("?" * len(period_mos))
        inc_rows  = conn.execute(f"SELECT SUM(amount) FROM bookings WHERE year=? AND month IN ({ph})",
                                 [exp_year]+period_mos).fetchone()[0] or 0
        opex_rows = conn.execute(f"SELECT SUM(amount) FROM manager_expenses WHERE year=? AND month IN ({ph})",
                                 [exp_year]+period_mos).fetchone()[0] or 0
        capex_rows_sum = conn.execute("SELECT SUM(amount) FROM capex_items WHERE "
                                      "CAST(substr(purchase_date,1,4) AS INTEGER)=?",
                                      (exp_year,)).fetchone()[0] or 0
        cohost_pct = float(get_setting("cohost_pct") or 30) / 100
        net_before = inc_rows - opex_rows
        cohost_share = net_before * cohost_pct
        owner_net    = net_before - cohost_share
        conn.close()

        st.divider()
        pm1,pm2,pm3,pm4 = st.columns(4)
        pm1.metric("Gross Income",   fmt_myr(inc_rows))
        pm2.metric("Total OpEx",     fmt_myr(opex_rows))
        pm3.metric("Total CapEx",    fmt_myr(capex_rows_sum))
        pm4.metric("Owner Net",      fmt_myr(owner_net))
        st.divider()

        dc1, dc2 = st.columns(2)
        with dc1:
            st.markdown("#### 📊 Excel Workbook")
            st.caption("Income Ledger · OpEx Breakdown · CapEx Schedule · P&L Summary · ROI")
            if st.button("Generate Excel", type="primary", key="gen_ulu_excel"):
                with st.spinner("Building workbook..."):
                    try:
                        import openpyxl as _xl
                        from openpyxl.styles import Font, PatternFill, Alignment
                        conn = get_db()

                        wb = _xl.Workbook()

                        # -- Cover --
                        ws0 = wb.active; ws0.title = "Cover"
                        ws0["A1"] = "ULU Mahsuri Villa — Financial Report"
                        ws0["A1"].font = Font(bold=True, size=14)
                        ws0["A2"] = f"Period: {exp_period} {exp_year}"
                        ws0["A3"] = f"Generated: {datetime.datetime.now().strftime('%d %b %Y %H:%M')}"
                        ws0["A5"] = "Contents:"
                        for i,s in enumerate(["Income Ledger","OpEx Breakdown","CapEx Schedule","P&L Summary"],1):
                            ws0[f"A{5+i}"] = f"  {i}. {s}"

                        # -- Income Ledger --
                        ws1 = wb.create_sheet("Income Ledger")
                        hdrs = ["Month","Guest","Type","Check-in","Check-out","Nights","Source","Amount (MYR)"]
                        for ci,h in enumerate(hdrs,1):
                            c = ws1.cell(1,ci,h); c.font=Font(bold=True,color="FFFFFF")
                            c.fill=PatternFill("solid",fgColor="2a3528")
                        bk_all = conn.execute(
                            f"SELECT * FROM bookings WHERE year=? AND month IN ({ph}) ORDER BY month,checkin",
                            [exp_year]+period_mos
                        ).fetchall()
                        for ri,r in enumerate(bk_all,2):
                            vals = [f"{MONTHS[r['month']-1]} {r['year']}", r["guest_name"],
                                    r["room_type"], r["checkin"], r["checkout"],
                                    r["nights"], r["source"], float(r["amount"] or 0)]
                            for ci,v in enumerate(vals,1):
                                ws1.cell(ri,ci,v)
                        ws1.cell(len(bk_all)+2,1,"TOTAL").font=Font(bold=True)
                        ws1.cell(len(bk_all)+2,8,sum(float(r["amount"] or 0) for r in bk_all)).font=Font(bold=True)

                        # -- OpEx Breakdown --
                        ws2 = wb.create_sheet("OpEx Breakdown")
                        hdrs2 = ["Month","Expense Item","Vendor","OpEx Category","Amount (MYR)"]
                        for ci,h in enumerate(hdrs2,1):
                            c = ws2.cell(1,ci,h); c.font=Font(bold=True,color="FFFFFF")
                            c.fill=PatternFill("solid",fgColor="2a3528")
                        ex_all = conn.execute(
                            f"SELECT * FROM manager_expenses WHERE year=? AND month IN ({ph}) ORDER BY month,id",
                            [exp_year]+period_mos
                        ).fetchall()
                        for ri,r in enumerate(ex_all,2):
                            vals2 = [f"{MONTHS[r['month']-1]} {r['year']}", r["expense_item"],
                                     r["vendor"], r["opex_category"] or "Uncategorised",
                                     float(r["amount"] or 0)]
                            for ci,v in enumerate(vals2,1):
                                ws2.cell(ri,ci,v)
                        ws2.cell(len(ex_all)+2,1,"TOTAL").font=Font(bold=True)
                        ws2.cell(len(ex_all)+2,5,sum(float(r["amount"] or 0) for r in ex_all)).font=Font(bold=True)

                        # -- CapEx Schedule --
                        ws3 = wb.create_sheet("CapEx Schedule")
                        hdrs3 = ["Date","Vendor","Description","Category","Amount (MYR)","Useful Life (yrs)","Annual Depreciation (MYR)"]
                        for ci,h in enumerate(hdrs3,1):
                            c = ws3.cell(1,ci,h); c.font=Font(bold=True,color="FFFFFF")
                            c.fill=PatternFill("solid",fgColor="C4856A")
                        cx_all = conn.execute(
                            "SELECT * FROM capex_items WHERE CAST(substr(purchase_date,1,4) AS INTEGER)=? ORDER BY purchase_date",
                            (exp_year,)
                        ).fetchall()
                        for ri,r in enumerate(cx_all,2):
                            amt = float(r["amount"] or 0)
                            life = int(r["useful_life_years"] or 5)
                            dep  = amt / life
                            vals3 = [r["purchase_date"],r["vendor"],r["description"],
                                     r["category"],amt,life,dep]
                            for ci,v in enumerate(vals3,1):
                                ws3.cell(ri,ci,v)
                        ws3.cell(len(cx_all)+2,1,"TOTAL").font=Font(bold=True)
                        ws3.cell(len(cx_all)+2,5,sum(float(r["amount"] or 0) for r in cx_all)).font=Font(bold=True)

                        # -- P&L Summary --
                        ws4 = wb.create_sheet("P&L Summary")
                        pnl = [
                            ("Gross Rental Income", inc_rows),
                            ("Less: Operating Expenses", -opex_rows),
                            ("Net Before Profit Sharing", net_before),
                            (f"Co-Host Share ({int(cohost_pct*100)}%)", -cohost_share),
                            ("Owner Net Profit", owner_net),
                            ("",""),
                            ("CapEx This Year (not in P&L)", capex_rows_sum),
                        ]
                        ws4["A1"]="P&L Summary"; ws4["A1"].font=Font(bold=True,size=12)
                        ws4["A2"]=f"{exp_period} {exp_year}"
                        for ri,(label,val) in enumerate(pnl,4):
                            ws4.cell(ri,1,label)
                            if val != "":
                                ws4.cell(ri,2,float(val))
                                if label in ("Owner Net Profit","Gross Rental Income"):
                                    ws4.cell(ri,1).font=Font(bold=True)
                                    ws4.cell(ri,2).font=Font(bold=True)

                        conn.close()
                        # Save
                        rpt_folder = os.path.join(os.path.dirname(os.path.abspath(__file__)),
                                                   "ULU Accountant Reports")
                        os.makedirs(rpt_folder, exist_ok=True)
                        fname_xl = f"ULU_Accounts_{exp_year}_{exp_period.split()[0]}.xlsx"
                        save_path = os.path.join(rpt_folder, fname_xl)
                        wb.save(save_path)
                        buf = io.BytesIO()
                        wb.save(buf); buf.seek(0)
                        st.success(f"✓ Saved to ULU Accountant Reports folder.")
                        st.download_button(f"⬇ {fname_xl}", data=buf.getvalue(),
                            file_name=fname_xl,
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                            key="dl_ulu_xl")
                    except Exception as e:
                        import traceback
                        st.error(f"Error: {e}")
                        st.code(traceback.format_exc())

        with dc2:
            st.markdown("#### 📄 PDF Financial Report")
            st.caption("A formatted PDF summary: P&L, bookings, expenses and CapEx — ready to send to your accountant.")
            if st.button("Generate PDF Report", type="primary", key="gen_ulu_pdf"):
                with st.spinner("Building PDF..."):
                    try:
                        conn = get_db()
                        bk_pdf = conn.execute(
                            f"SELECT * FROM bookings WHERE year=? AND month IN ({ph}) ORDER BY month,checkin",
                            [exp_year]+period_mos
                        ).fetchall()
                        ex_pdf = conn.execute(
                            f"SELECT * FROM manager_expenses WHERE year=? AND month IN ({ph}) ORDER BY month,id",
                            [exp_year]+period_mos
                        ).fetchall()
                        pe_pdf = conn.execute(
                            f"SELECT * FROM personal_expenses WHERE year=? AND month IN ({ph}) ORDER BY month,id",
                            [exp_year]+period_mos
                        ).fetchall()
                        cx_pdf = conn.execute(
                            "SELECT * FROM capex_items WHERE CAST(substr(purchase_date,1,4) AS INTEGER)=? ORDER BY purchase_date",
                            (exp_year,)
                        ).fetchall()
                        conn.close()

                        buf_pdf = io.BytesIO()
                        doc = SimpleDocTemplate(buf_pdf, pagesize=A4,
                            rightMargin=18*mm, leftMargin=18*mm,
                            topMargin=16*mm, bottomMargin=16*mm)

                        INK   = colors.HexColor("#1C1C1A")
                        GREEN = colors.HexColor("#2a3528")
                        CREAM = colors.HexColor("#F5F0E8")
                        ACCENT= colors.HexColor("#C4856A")
                        LIGHT = colors.HexColor("#E5DDD0")
                        GREY  = colors.HexColor("#6B6560")
                        W = A4[0] - 36*mm

                        sN = ParagraphStyle("n", fontName="Helvetica", fontSize=9, leading=13, textColor=INK)
                        sS = ParagraphStyle("s", fontName="Helvetica", fontSize=8, leading=11, textColor=GREY)
                        sB = ParagraphStyle("b", fontName="Helvetica-Bold", fontSize=9, leading=13, textColor=INK)
                        sT = ParagraphStyle("t", fontName="Helvetica-Bold", fontSize=16, leading=20, textColor=INK)
                        sSub = ParagraphStyle("su", fontName="Helvetica-Bold", fontSize=11, leading=14, textColor=GREEN)
                        sR  = ParagraphStyle("r", fontName="Helvetica", fontSize=9, alignment=TA_RIGHT, textColor=INK)
                        sRB = ParagraphStyle("rb", fontName="Helvetica-Bold", fontSize=10, alignment=TA_RIGHT, textColor=ACCENT)

                        def th(txt): return Paragraph(txt, ParagraphStyle("th", fontName="Helvetica-Bold", fontSize=8, textColor=colors.white))
                        def td(txt): return Paragraph(str(txt), sS)
                        def td_r(txt): return Paragraph(str(txt), ParagraphStyle("tdr", fontName="Helvetica", fontSize=8, alignment=TA_RIGHT, textColor=INK))

                        story = []

                        # Header
                        story.append(Paragraph("ULU Mahsuri Villa", sT))
                        story.append(Paragraph(f"Financial Report — {exp_period} {exp_year}", sSub))
                        story.append(Paragraph(f"Generated: {datetime.datetime.now().strftime('%d %B %Y %H:%M')}", sS))
                        story.append(HRFlowable(width="100%", thickness=2, color=GREEN, spaceAfter=14))

                        # P&L Summary box
                        story.append(Paragraph("1. P&L Summary", sSub))
                        story.append(Spacer(1, 6))
                        pnl_data = [
                            ["Description", "Amount (RM)"],
                            ["Gross Rental Income", f"{inc_rows:,.2f}"],
                            ["  Manager Expenses", f"({opex_rows:,.2f})"],
                            ["  Net Before Sharing", f"{net_before:,.2f}"],
                            [f"  Co-Host Share ({int(cohost_pct*100)}%)", f"({cohost_share:,.2f})"],
                            ["Owner Net Profit", f"{owner_net:,.2f}"],
                            ["", ""],
                            ["CapEx This Year (memo only)", f"{capex_rows_sum:,.2f}"],
                        ]
                        pnl_tbl = Table(pnl_data, colWidths=[W*0.7, W*0.3])
                        pnl_tbl.setStyle(TableStyle([
                            ("BACKGROUND",(0,0),(-1,0), GREEN),
                            ("TEXTCOLOR",(0,0),(-1,0), colors.white),
                            ("FONTNAME",(0,0),(-1,0),"Helvetica-Bold"),
                            ("FONTNAME",(0,5),(-1,5),"Helvetica-Bold"),
                            ("FONTSIZE",(0,0),(-1,-1),9),
                            ("ALIGN",(1,0),(-1,-1),"RIGHT"),
                            ("LINEBELOW",(0,3),(-1,3),0.5,LIGHT),
                            ("LINEABOVE",(0,5),(-1,5),1,GREEN),
                            ("TOPPADDING",(0,0),(-1,-1),5),
                            ("BOTTOMPADDING",(0,0),(-1,-1),5),
                            ("LEFTPADDING",(0,0),(-1,-1),8),
                        ]))
                        story.append(pnl_tbl)
                        story.append(Spacer(1, 16))

                        # Bookings
                        story.append(Paragraph("2. Booking Income", sSub))
                        story.append(Spacer(1, 6))
                        if bk_pdf:
                            bk_head = [th("Month"), th("Guest"), th("Type"), th("Check-in"), th("Check-out"), th("Nts"), th("Src"), th("Amount (RM)")]
                            bk_rows_tbl = [bk_head]
                            for b in bk_pdf:
                                bk_rows_tbl.append([
                                    td(MONTHS[b["month"]-1]),
                                    td(b["guest_name"] or ""),
                                    td(b["room_type"] or ""),
                                    td(b["checkin"] or ""),
                                    td(b["checkout"] or ""),
                                    td(str(b["nights"] or "")),
                                    td(b["source"] or ""),
                                    td_r(f"{float(b['amount'] or 0):,.2f}")
                                ])
                            bk_tbl = Table(bk_rows_tbl, colWidths=[W*0.09,W*0.17,W*0.07,W*0.11,W*0.11,W*0.05,W*0.07,W*0.13])
                            bk_tbl.setStyle(TableStyle([
                                ("BACKGROUND",(0,0),(-1,0),GREEN),
                                ("ROWBACKGROUNDS",(0,1),(-1,-1),[colors.white,colors.HexColor("#FAF8F5")]),
                                ("TOPPADDING",(0,0),(-1,-1),4),("BOTTOMPADDING",(0,0),(-1,-1),4),
                                ("LEFTPADDING",(0,0),(-1,-1),5),("RIGHTPADDING",(0,0),(-1,-1),5),
                                ("LINEBELOW",(0,0),(-1,-1),0.3,LIGHT),
                            ]))
                            story.append(bk_tbl)
                            story.append(Paragraph(f"Total: RM {sum(float(b['amount'] or 0) for b in bk_pdf):,.2f}", sB))
                        else:
                            story.append(Paragraph("No bookings for this period.", sS))
                        story.append(Spacer(1, 14))

                        # Manager expenses
                        story.append(Paragraph("3. Manager's Operating Expenses", sSub))
                        story.append(Spacer(1, 6))
                        if ex_pdf:
                            ex_head = [th("Month"), th("Expense Item"), th("Vendor"), th("Category"), th("Amount (RM)")]
                            ex_rows_tbl = [ex_head]
                            for e in ex_pdf:
                                ex_rows_tbl.append([
                                    td(MONTHS[e["month"]-1]),
                                    td(e["expense_item"] or ""),
                                    td(e["vendor"] or ""),
                                    td(e["opex_category"] or "Uncategorised"),
                                    td_r(f"{float(e['amount'] or 0):,.2f}")
                                ])
                            ex_tbl = Table(ex_rows_tbl, colWidths=[W*0.1,W*0.22,W*0.22,W*0.28,W*0.18])
                            ex_tbl.setStyle(TableStyle([
                                ("BACKGROUND",(0,0),(-1,0),GREEN),
                                ("ROWBACKGROUNDS",(0,1),(-1,-1),[colors.white,colors.HexColor("#FAF8F5")]),
                                ("TOPPADDING",(0,0),(-1,-1),4),("BOTTOMPADDING",(0,0),(-1,-1),4),
                                ("LEFTPADDING",(0,0),(-1,-1),5),("RIGHTPADDING",(0,0),(-1,-1),5),
                                ("LINEBELOW",(0,0),(-1,-1),0.3,LIGHT),
                            ]))
                            story.append(ex_tbl)
                            story.append(Paragraph(f"Total: RM {sum(float(e['amount'] or 0) for e in ex_pdf):,.2f}", sB))
                        else:
                            story.append(Paragraph("No manager expenses for this period.", sS))
                        story.append(Spacer(1, 14))

                        # Personal expenses
                        story.append(Paragraph("4. Owner Personal Expenses (ULU Share)", sSub))
                        story.append(Spacer(1, 6))
                        if pe_pdf:
                            pe_head = [th("Month"), th("Vendor"), th("Category"), th("Description"), th("ULU Share (RM)")]
                            pe_rows_tbl = [pe_head]
                            for p in pe_pdf:
                                pe_rows_tbl.append([
                                    td(MONTHS[p["month"]-1]),
                                    td(p["vendor"] or ""),
                                    td(p["category"] or ""),
                                    td((p["description"] or "")[:40]),
                                    td_r(f"{float(p['ulu_share'] or 0):,.2f}")
                                ])
                            pe_tbl = Table(pe_rows_tbl, colWidths=[W*0.1,W*0.2,W*0.2,W*0.32,W*0.18])
                            pe_tbl.setStyle(TableStyle([
                                ("BACKGROUND",(0,0),(-1,0),GREEN),
                                ("ROWBACKGROUNDS",(0,1),(-1,-1),[colors.white,colors.HexColor("#FAF8F5")]),
                                ("TOPPADDING",(0,0),(-1,-1),4),("BOTTOMPADDING",(0,0),(-1,-1),4),
                                ("LEFTPADDING",(0,0),(-1,-1),5),("RIGHTPADDING",(0,0),(-1,-1),5),
                                ("LINEBELOW",(0,0),(-1,-1),0.3,LIGHT),
                            ]))
                            story.append(pe_tbl)
                            story.append(Paragraph(f"Total ULU Share: RM {sum(float(p['ulu_share'] or 0) for p in pe_pdf):,.2f}", sB))
                        else:
                            story.append(Paragraph("No personal expenses for this period.", sS))
                        story.append(Spacer(1, 14))

                        # CapEx Schedule
                        story.append(Paragraph("5. CapEx Schedule", sSub))
                        story.append(Spacer(1, 6))
                        if cx_pdf:
                            cx_head = [th("Date"), th("Vendor"), th("Description"), th("Category"), th("Amount (RM)"), th("Life"), th("Dep/yr")]
                            cx_rows_tbl = [cx_head]
                            for c in cx_pdf:
                                amt = float(c["amount"] or 0)
                                life = int(c["useful_life_years"] or 5)
                                cx_rows_tbl.append([
                                    td(c["purchase_date"] or ""),
                                    td(c["vendor"] or ""),
                                    td((c["description"] or "")[:30]),
                                    td(c["category"] or ""),
                                    td_r(f"{amt:,.2f}"),
                                    td_r(str(life)),
                                    td_r(f"{amt/life:,.2f}")
                                ])
                            cx_tbl = Table(cx_rows_tbl, colWidths=[W*0.11,W*0.15,W*0.22,W*0.2,W*0.12,W*0.07,W*0.13])
                            cx_tbl.setStyle(TableStyle([
                                ("BACKGROUND",(0,0),(-1,0),ACCENT),
                                ("ROWBACKGROUNDS",(0,1),(-1,-1),[colors.white,colors.HexColor("#FAF8F5")]),
                                ("TOPPADDING",(0,0),(-1,-1),4),("BOTTOMPADDING",(0,0),(-1,-1),4),
                                ("LEFTPADDING",(0,0),(-1,-1),5),("RIGHTPADDING",(0,0),(-1,-1),5),
                                ("LINEBELOW",(0,0),(-1,-1),0.3,LIGHT),
                            ]))
                            story.append(cx_tbl)
                            story.append(Paragraph(f"Total CapEx: RM {sum(float(c['amount'] or 0) for c in cx_pdf):,.2f}", sB))
                        else:
                            story.append(Paragraph("No CapEx items for this year.", sS))

                        story.append(Spacer(1, 18))
                        story.append(HRFlowable(width="100%", thickness=0.5, color=LIGHT, spaceAfter=8))
                        story.append(Paragraph(
                            f"ULU Mahsuri Villa · {exp_period} {exp_year} · Generated {datetime.datetime.now().strftime('%d %B %Y')}",
                            sS
                        ))

                        doc.build(story)
                        buf_pdf.seek(0)
                        fname_pdf = f"ULU_FinancialReport_{exp_year}_{exp_period.split()[0]}.pdf"
                        # Save to reports folder
                        rpt_folder = os.path.join(os.path.dirname(os.path.abspath(__file__)), "ULU Accountant Reports")
                        os.makedirs(rpt_folder, exist_ok=True)
                        with open(os.path.join(rpt_folder, fname_pdf), "wb") as _f:
                            _f.write(buf_pdf.getvalue())
                        st.success(f"✓ PDF saved to ULU Accountant Reports folder.")
                        st.download_button(f"⬇ {fname_pdf}", data=buf_pdf.getvalue(),
                            file_name=fname_pdf, mime="application/pdf", key="dl_ulu_pdf")
                    except Exception as e:
                        import traceback
                        st.error(f"Error: {e}")
                        st.code(traceback.format_exc())

            st.divider()
            st.markdown("#### 📁 Supporting Documents")
            st.caption("Share these folders with your accountant via OneDrive for full audit trail.")
            base_dir = os.path.dirname(os.path.abspath(__file__))
            folders = {
                "📋 Manager Monthly Reports": "Manager Monthly Reports",
                "📋 CapEx Receipts": "CapEx Receipts",
                "📊 Accountant Reports": "ULU Accountant Reports",
            }
            for label, folder in folders.items():
                fpath = os.path.join(base_dir, folder)
                if os.path.exists(fpath):
                    file_count = sum(len(files) for _, _, files in os.walk(fpath))
                    st.write(f"{label} — {file_count} file(s)")
                else:
                    st.write(f"{label} — folder not yet created")

    with acct_tab2:
        st.subheader("Share Access with Accountant / Tax Agent")
        share1, share2 = st.tabs(["☁️ OneDrive Links", "📋 What to Share"])

        with share1:
            st.caption("Paste your OneDrive sharing links below to generate a WhatsApp message.")
            od_mgr     = st.text_input("📋 Manager Monthly Reports folder link", key="od_ulu_mgr",
                placeholder="Right-click folder → Share → Copy link")
            od_capex   = st.text_input("📋 CapEx Receipts folder link", key="od_ulu_capex",
                placeholder="Right-click folder → Share → Copy link")
            od_reports = st.text_input("📊 ULU Accountant Reports folder link", key="od_ulu_rpt",
                placeholder="Right-click folder → Share → Copy link")

            if st.button("📱 Generate WhatsApp Message", key="gen_ulu_wa"):
                links = ""
                if od_reports: links += f"📊 *Financial Reports (Excel):*\n{od_reports}\n\n"
                if od_mgr:     links += f"📋 *Manager Monthly Reports (Azary's Submissions):*\n{od_mgr}\n\n"
                if od_capex:   links += f"📋 *CapEx Receipts & Invoices:*\n{od_capex}\n\n"
                if not links:
                    st.warning("Please paste at least one link.")
                else:
                    wa = (
                        f"Hi, please find below the shared folders for "
                        f"ULU Mahsuri Villa financial documents.\n\n"
                        f"{links}"
                        f"All folders are view-only. "
                        f"Please contact me if you need any assistance.\n"
                        f"— Azlan"
                    )
                    st.text_area("📱 WhatsApp Message (copy & paste):", value=wa,
                                 height=250, key="ulu_wa_msg")
                    st.success("✓ Message ready. Copy and send via WhatsApp.")

        with share2:
            st.markdown("""
**What to prepare for your accountant / tax agent:**

**For Income Tax (LHDN):**
- ✅ Excel Workbook (Income Ledger + P&L Summary)
- ✅ Manager Monthly Reports (Azary's original billing submissions)
- ✅ Airbnb statements for the year
- ✅ CapEx Schedule (for depreciation claims)

**For RPGT (if property sold):**
- ✅ CapEx receipts (all upgrade costs increase your cost base)
- ✅ Original purchase documents

**CapEx vs OpEx for tax:**
- **OpEx** — fully deductible against rental income in the same year
- **CapEx** — depreciated over useful life (typically 5-10 years)
- CapEx receipts must be kept as evidence for 7 years

**What the accountant needs each year:**
1. Generate Excel report from this tab
2. Share CapEx Receipts folder (all original invoices)
3. Share ULU Accountant Reports folder
""")
    st.markdown('</div>', unsafe_allow_html=True)


# ─────────────────────────────────────────────
# HIDDEN IMPORT TOOL (run once via URL param)
# Access at: localhost:8501/?import=true
# ─────────────────────────────────────────────
import_mode = st.query_params.get("import", "false")
if import_mode == "true":
    st.divider()
    st.markdown("## 📥 Bulk Data Import")
    st.info("Upload the CSV files provided to load historical data in one go.")

    col_i1, col_i2 = st.columns(2)

    with col_i1:
        bk_file = st.file_uploader("Upload Bookings CSV", type=["csv"], key="imp_bk")
        if bk_file and st.button("Import Bookings"):
            df_bk = pd.read_csv(bk_file)
            conn = get_db()
            imported = 0
            for _, row in df_bk.iterrows():
                conn.execute(
                    "INSERT INTO bookings (year,month,guest_name,room_type,checkin,checkout,nights,source,amount,notes) VALUES (?,?,?,?,?,?,?,?,?,?)",
                    (int(row["year"]), int(row["month"]), row["guest_name"], row["room_type"],
                     row["checkin"], row["checkout"], int(row["nights"]), row["source"],
                     float(row["amount"]), str(row.get("notes","")))
                )
                imported += 1
            conn.commit(); conn.close()
            st.success(f"✅ Imported {imported} bookings!")

    with col_i2:
        ex_file = st.file_uploader("Upload Manager Expenses CSV", type=["csv"], key="imp_ex")
        if ex_file and st.button("Import Expenses"):
            df_ex = pd.read_csv(ex_file)
            conn = get_db()
            imported = 0
            for _, row in df_ex.iterrows():
                conn.execute(
                    "INSERT INTO manager_expenses (year,month,expense_item,vendor,amount,notes) VALUES (?,?,?,?,?,?)",
                    (int(row["year"]), int(row["month"]), row["expense_item"], row["vendor"],
                     float(row["amount"]), str(row.get("notes","")))
                )
                imported += 1
            conn.commit(); conn.close()
            st.success(f"✅ Imported {imported} expense lines!")

    st.divider()
    pe_file = st.file_uploader("Upload Personal Expenses CSV", type=["csv"], key="imp_pe")
    if pe_file and st.button("Import Personal Expenses"):
        df_pe = pd.read_csv(pe_file)
        conn = get_db()
        imported = 0
        for _, row in df_pe.iterrows():
            conn.execute(
                "INSERT INTO personal_expenses (year,month,vendor,bill_date,category,description,total_amount,ulu_share,file_name) VALUES (?,?,?,?,?,?,?,?,?)",
                (int(row["year"]), int(row["month"]), row["vendor"], row["bill_date"],
                 row["category"], row["description"], float(row["total_amount"]),
                 float(row["ulu_share"]), str(row.get("file_name","")))
            )
            imported += 1
        conn.commit(); conn.close()
        st.success(f"✅ Imported {imported} personal expense entries!")

    st.warning("⚠️ Only import once! Importing again will create duplicate entries.")

# ══════════════════════════════════════════════
# ══════════════════════════════════════════════
# TAB 9 — PAYMENTS & VOUCHERS
# ══════════════════════════════════════════════
PAYMENT_STATUSES = ["Pending", "Paid", "Partial", "Cancelled"]
PAYMENT_TYPES    = [
    "Property Manager Reimbursable",
    "Property Management Fee (Archmedia Sdn Bhd)",
    "Supplier / Vendor Payment",
    "Utility Bill",
    "Housekeeping & Laundry",
    "Pool Maintenance",
    "Maintenance & Repairs",
    "Contractor Payment",
    "Other",
]
STATUS_ICONS = {"Pending":"🟡","Paid":"🟢","Partial":"🔵","Cancelled":"⚫"}

def generate_voucher_number(year, month):
    conn = get_db()
    rows = conn.execute("SELECT id FROM payments WHERE year=? AND month=? ORDER BY id",(year,month)).fetchall()
    conn.close()
    return f"PV-{year}-{month:02d}-{len(rows)+1:03d}"

def generate_payment_voucher_pdf(payment: dict) -> bytes:
    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf,pagesize=A4,rightMargin=20*mm,leftMargin=20*mm,topMargin=18*mm,bottomMargin=18*mm)
    INK=colors.HexColor("#1C1C1A"); GREEN=colors.HexColor("#2a3528"); LIGHT=colors.HexColor("#E5DDD0"); RED=colors.HexColor("#CC0000")
    W=A4[0]-40*mm
    sN =ParagraphStyle("n", fontName="Helvetica",     fontSize=10,leading=14,textColor=INK)
    sB =ParagraphStyle("b", fontName="Helvetica-Bold",fontSize=10,leading=14,textColor=INK)
    sS =ParagraphStyle("s", fontName="Helvetica",     fontSize=8, leading=11,textColor=colors.HexColor("#6B6560"))
    sT =ParagraphStyle("t", fontName="Helvetica-Bold",fontSize=18,leading=22,textColor=INK)
    sSub=ParagraphStyle("su",fontName="Helvetica-Bold",fontSize=11,leading=14,textColor=GREEN)
    sCan=ParagraphStyle("c", fontName="Helvetica-Bold",fontSize=22,alignment=TA_CENTER,textColor=RED)
    sCanR=ParagraphStyle("cr",fontName="Helvetica-Bold",fontSize=11,alignment=TA_CENTER,textColor=RED)
    is_cancelled  = payment.get("status","")=="Cancelled"
    cancel_reason = payment.get("cancellation_reason","") or ""
    month_val=payment.get("month",1)
    try: month_label=MONTHS[int(month_val)-1]
    except: month_label=str(month_val)
    payee  = payment.get("payee_name","") or payment.get("payee","") or "—"
    ptype  = payment.get("payee_type","") or payment.get("payment_type","") or "—"
    amount = float(payment.get("amount_due",0) or payment.get("amount",0) or 0)
    ref    = payment.get("reference_no","") or payment.get("payment_ref","") or "—"
    bank   = payment.get("payee_bank","") or "—"
    acc    = payment.get("payee_account","") or "—"
    voucher= payment.get("voucher_no","") or f"PMT-{payment.get('id','?')}"
    pdate  = payment.get("payment_date","") or "—"
    method = payment.get("payment_method","") or "—"
    status = payment.get("status","") or "—"
    desc   = payment.get("description","") or "—"
    story=[]
    story.append(Paragraph("ULU Mahsuri Villa",sT))
    story.append(Paragraph("Payment Voucher",sSub))
    story.append(HRFlowable(width="100%",thickness=2,color=GREEN,spaceAfter=10))
    if is_cancelled:
        story.append(Paragraph("CANCELLED",sCan))
        if cancel_reason: story.append(Paragraph(f"Reason: {cancel_reason}",sCanR))
        story.append(Spacer(1,8))
        story.append(HRFlowable(width="100%",thickness=1,color=RED,spaceAfter=8))
    details=[
        ["Voucher No.",voucher,  "Date",   str(pdate)],
        ["Payee",      payee,    "Month",  f"{month_label} {payment.get('year','—')}"],
        ["Type",       ptype,    "Method", method],
        ["Bank",       bank,     "Acc No.",acc],
        ["Reference",  ref,      "Status", status],
    ]
    det_tbl=Table(details,colWidths=[W*0.18,W*0.32,W*0.18,W*0.32])
    det_tbl.setStyle(TableStyle([
        ("FONTNAME",(0,0),(-1,-1),"Helvetica"),("FONTNAME",(0,0),(0,-1),"Helvetica-Bold"),
        ("FONTNAME",(2,0),(2,-1),"Helvetica-Bold"),("FONTSIZE",(0,0),(-1,-1),9),
        ("TEXTCOLOR",(0,0),(-1,-1),INK),("TOPPADDING",(0,0),(-1,-1),5),
        ("BOTTOMPADDING",(0,0),(-1,-1),5),("LINEBELOW",(0,0),(-1,-1),0.3,LIGHT),
    ]))
    story.append(det_tbl); story.append(Spacer(1,12))
    story.append(Paragraph("Description / Purpose:",sB)); story.append(Paragraph(desc,sN)); story.append(Spacer(1,12))
    bg_col=colors.HexColor("#888") if is_cancelled else GREEN
    amt_tbl=Table([["Amount (MYR)",f"RM {amount:,.2f}"]],colWidths=[W*0.6,W*0.4])
    amt_tbl.setStyle(TableStyle([
        ("BACKGROUND",(0,0),(-1,-1),bg_col),("TEXTCOLOR",(0,0),(-1,-1),colors.white),
        ("FONTNAME",(0,0),(-1,-1),"Helvetica-Bold"),("FONTSIZE",(0,0),(-1,-1),13),
        ("ALIGN",(1,0),(1,0),"RIGHT"),("TOPPADDING",(0,0),(-1,-1),10),
        ("BOTTOMPADDING",(0,0),(-1,-1),10),("LEFTPADDING",(0,0),(-1,-1),12),("RIGHTPADDING",(0,0),(-1,-1),12),
    ]))
    story.append(amt_tbl)
    if is_cancelled and cancel_reason:
        story.append(Spacer(1,6)); story.append(Paragraph(f"Cancellation reason: {cancel_reason}",sCanR))
    story.append(Spacer(1,16))
    story.append(HRFlowable(width="100%",thickness=0.5,color=LIGHT,spaceAfter=8))
    story.append(Paragraph(f"ULU Mahsuri Villa · Generated {datetime.datetime.now().strftime('%d %B %Y %H:%M')}",sS))
    doc.build(story); buf.seek(0); return buf.getvalue()

def pv_year_val(r):
    y=r.get("year")
    if y: return int(y)
    ca=str(r.get("created_at",""))
    return int(ca[:4]) if len(ca)>=4 else 0

def pv_month_val(r):
    m=r.get("month")
    if m: return int(m)
    ca=str(r.get("created_at",""))
    try: return int(ca[5:7])
    except: return 0

with tab9:
    st.markdown('<div class="card">', unsafe_allow_html=True)
    st.markdown('<p class="card-title">💳 Payments & Vouchers</p>', unsafe_allow_html=True)
    st.markdown("Record payments to vendors, Property Manager reimbursables, and Archmedia Sdn Bhd management fees. Generate PDF payment vouchers for each transaction.")

    pv_subtab1, pv_subtab2 = st.tabs(["➕ New Payment", "📋 All Payment Records"])

    with pv_subtab1:
        ym_list_pv = get_year_month_list()
        ym_labels_pv = [f"{MONTHS[m-1]} {y} ({operation_year(y,m)})" for y,m in ym_list_pv]
        sel_idx_pv = st.selectbox("Assign payment to month", range(len(ym_labels_pv)),
            format_func=lambda i: ym_labels_pv[i], key="pv_month")
        pv_year, pv_month = ym_list_pv[sel_idx_pv]

        col_a, col_b = st.columns(2, gap="large")
        with col_a:
            pv_payee  = st.text_input("Payee Name", placeholder="e.g. Archmedia Sdn Bhd", key="pv_payee")
            pv_type   = st.selectbox("Payment Type", PAYMENT_TYPES, key="pv_type")
            pv_desc   = st.text_area("Description / Purpose", height=80, key="pv_desc",
                            placeholder="e.g. Property Management Fee for June 2026")
            pv_amount = st.number_input("Amount (RM)", min_value=0.0, step=10.0, format="%.2f", key="pv_amount")
            pv_date   = st.text_input("Payment Date (YYYY-MM-DD)", value=datetime.date.today().isoformat(), key="pv_date")
        with col_b:
            pv_method = st.selectbox("Payment Method",
                ["Bank Transfer","Cash","Cheque","Online Banking","Other"], key="pv_method")
            pv_ref    = st.text_input("Payment Reference / Transaction ID",
                            placeholder="e.g. IBG20260705001", key="pv_ref")
            pv_bank   = st.text_input("Payee Bank", placeholder="e.g. Maybank / CIMB / RHB", key="pv_bank")
            pv_acc    = st.text_input("Payee Account No.", placeholder="e.g. 1234567890", key="pv_acc")
            pv_notes  = st.text_area("Notes", height=60, key="pv_notes")

        if st.button("💾 Save Payment & Generate Voucher", type="primary", key="btn_save_pv"):
            if not pv_payee or pv_amount <= 0:
                st.error("Payee name and amount are required.")
            else:
                voucher_no = generate_voucher_number(pv_year, pv_month)
                conn = get_db()
                conn.execute(
                    """INSERT INTO payments
                       (year,month,voucher_no,payee_name,payee_type,description,amount_due,
                        payment_date,payment_method,reference_no,payee_bank,payee_account,
                        notes,status)
                       VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?)""",
                    (pv_year,pv_month,voucher_no,pv_payee,pv_type,pv_desc,
                     pv_amount,pv_date,pv_method,pv_ref,pv_bank,pv_acc,pv_notes,"Pending")
                )
                conn.commit(); conn.close()
                pd_dict = {
                    "voucher_no":voucher_no,"year":pv_year,"month":pv_month,
                    "payee_name":pv_payee,"payee_type":pv_type,"description":pv_desc,
                    "amount_due":pv_amount,"payment_date":pv_date,"payment_method":pv_method,
                    "reference_no":pv_ref,"payee_bank":pv_bank,"payee_account":pv_acc,
                    "status":"Pending","cancellation_reason":""
                }
                pdf_bytes = generate_payment_voucher_pdf(pd_dict)
                st.success(f"Payment saved. Voucher: **{voucher_no}**")
                st.download_button(f"⬇️ Download {voucher_no}.pdf", data=pdf_bytes,
                    file_name=f"{voucher_no}.pdf", mime="application/pdf", key=f"dl_new_{voucher_no}")
                st.rerun()

    with pv_subtab2:
        pf1, pf2, pf3 = st.columns(3)
        filter_year   = pf1.selectbox("Year",  list(range(datetime.datetime.now().year, 2023, -1)), key="pv_filter_yr")
        filter_month  = pf2.selectbox("Month", ["All"] + MONTHS, key="pv_filter_mo")
        filter_status = pf3.selectbox("Status", ["All"] + PAYMENT_STATUSES, key="pv_filter_st")

        conn = get_db()
        all_pv = conn.execute("SELECT * FROM payments ORDER BY id DESC", ()).fetchall()
        conn.close()

        pv_rows = [r for r in all_pv if pv_year_val(r) == filter_year]
        if filter_month != "All":
            mo_num = MONTHS.index(filter_month) + 1
            pv_rows = [r for r in pv_rows if pv_month_val(r) == mo_num]
        if filter_status != "All":
            pv_rows = [r for r in pv_rows if r.get("status","") == filter_status]

        active_rows  = [r for r in pv_rows if r.get("status","") != "Cancelled"]
        total_amount = sum(float(r.get("amount_due",0) or 0) for r in active_rows)
        paid_rows    = [r for r in active_rows if r.get("status","") == "Paid"]
        total_paid   = sum(float(r.get("amount_due",0) or 0) for r in paid_rows)
        total_pending= total_amount - total_paid

        mc1,mc2,mc3,mc4 = st.columns(4)
        mc1.metric("Total Records", len(pv_rows))
        mc2.metric("Total Due (excl. cancelled)", fmt_myr(total_amount))
        mc3.metric("Total Paid", fmt_myr(total_paid))
        mc4.metric("Outstanding", fmt_myr(total_pending))
        st.divider()

        if not pv_rows:
            st.info("No payment records for this period.")
        else:
            for r in [dict(r) for r in pv_rows]:
                r_id        = r.get("id")
                r_status    = r.get("status","Pending") or "Pending"
                r_icon      = STATUS_ICONS.get(r_status,"⚪")
                is_canc     = r_status == "Cancelled"
                canc_reason = r.get("cancellation_reason","") or ""
                payee_disp  = r.get("payee_name","") or r.get("payee","") or "—"
                amount_disp = float(r.get("amount_due",0) or 0)
                voucher_disp= r.get("voucher_no","") or f"PMT-{r_id}"
                mo_d = pv_month_val(r); yr_d = pv_year_val(r)

                header_label = f"{r_icon} **{voucher_disp}** — {payee_disp} — {fmt_myr(amount_disp)}"
                if is_canc: header_label += " *(Cancelled)*"

                with st.expander(header_label, expanded=False):
                    if is_canc and canc_reason:
                        st.warning(f"🚫 Cancelled — Reason: {canc_reason}")

                    d1,d2,d3 = st.columns(3)
                    d1.markdown(f"**Date:** {r.get('payment_date','—')}")
                    d2.markdown(f"**Month:** {MONTHS[mo_d-1] if mo_d else '—'} {yr_d or '—'}")
                    d3.markdown(f"**Type:** {r.get('payee_type','') or r.get('payment_type','') or '—'}")
                    d4,d5,d6 = st.columns(3)
                    d4.markdown(f"**Method:** {r.get('payment_method','—')}")
                    d5.markdown(f"**Ref:** {r.get('reference_no','') or r.get('payment_ref','') or '—'}")
                    d6.markdown(f"**Bank:** {r.get('payee_bank','—')} / {r.get('payee_account','—')}")
                    st.markdown(f"**Description:** {r.get('description','—')}")
                    if r.get("notes"): st.markdown(f"**Notes:** {r.get('notes','')}")

                    st.divider()
                    act1,act2,act3,act4 = st.columns([2,1,2,2])

                    if not is_canc:
                        safe_statuses = [s for s in PAYMENT_STATUSES if s != "Cancelled"]
                        cur_idx = safe_statuses.index(r_status) if r_status in safe_statuses else 0
                        new_status = act1.selectbox("Status", safe_statuses, index=cur_idx,
                            key=f"pv_status_{r_id}", label_visibility="collapsed")
                        if act2.button("Update", key=f"pv_upd_{r_id}"):
                            conn = get_db()
                            conn.execute("UPDATE payments SET status=? WHERE id=?",(new_status,r_id))
                            conn.commit(); conn.close(); st.rerun()
                    else:
                        act1.markdown("<span class='badge-cancelled'>⚫ Cancelled</span>", unsafe_allow_html=True)

                    pdf_bytes = generate_payment_voucher_pdf(r)
                    suffix = "_CANCELLED" if is_canc else ""
                    act3.download_button("⬇️ Voucher PDF", data=pdf_bytes,
                        file_name=f"{voucher_disp}{suffix}.pdf", mime="application/pdf",
                        key=f"dl_pv_{r_id}")

                    if not is_canc:
                        if act4.button("🚫 Cancel Record", key=f"pv_cancel_btn_{r_id}"):
                            st.session_state[f"show_cancel_{r_id}"] = True
                    else:
                        if act4.button("🗑️ Delete", key=f"pv_del_{r_id}"):
                            conn = get_db()
                            conn.execute("DELETE FROM payments WHERE id=?",(r_id,))
                            conn.commit(); conn.close(); st.rerun()

                    if st.session_state.get(f"show_cancel_{r_id}"):
                        st.markdown("---")
                        st.error("**Cancel this payment record — this cannot be undone.**")
                        cancel_input = st.text_input("Reason for cancellation (required)",
                            key=f"cancel_reason_{r_id}",
                            placeholder="e.g. Duplicate entry / Wrong amount / Payment reversed")
                        cc1,cc2 = st.columns(2)
                        if cc1.button("Confirm Cancellation", key=f"cancel_confirm_{r_id}", type="primary"):
                            if not cancel_input.strip():
                                st.error("A reason is required before confirming cancellation.")
                            else:
                                conn = get_db()
                                conn.execute(
                                    "UPDATE payments SET status='Cancelled', cancellation_reason=? WHERE id=?",
                                    (cancel_input.strip(), r_id)
                                )
                                conn.commit(); conn.close()
                                st.session_state.pop(f"show_cancel_{r_id}", None)
                                st.success("Record cancelled and reason recorded.")
                                st.rerun()
                        if cc2.button("Back", key=f"cancel_back_{r_id}"):
                            st.session_state.pop(f"show_cancel_{r_id}", None); st.rerun()

    st.markdown('</div>', unsafe_allow_html=True)


# ══════════════════════════════════════════════
# TAB 10 — DIRECT INCOME & EXTRAS
# ══════════════════════════════════════════════
DIRECT_INCOME_TYPES = [
    "Direct Booking (Cash/Transfer)",
    "Extra Mattress",
    "Baby Cot",
    "Extra Cleaning",
    "Damage Compensation",
    "Cancellation Fee",
    "Other Extra Charge",
]

def parse_airbnb_csv(file_bytes: bytes) -> list:
    """Parse Airbnb CSV export and return list of reservation dicts."""
    import csv, io
    text = file_bytes.decode("utf-8-sig")
    reader = csv.DictReader(io.StringIO(text))
    reservations = []
    for row in reader:
        if row.get("Type","").strip() == "Reservation":
            try:
                reservations.append({
                    "confirmation": row.get("Confirmation code","").strip(),
                    "guest":        row.get("Guest","").strip(),
                    "checkin":      row.get("Start date","").strip(),
                    "checkout":     row.get("End date","").strip(),
                    "nights":       int(float(row.get("Nights",0) or 0)),
                    "payout":       float(row.get("Amount",0) or 0),
                    "gross":        float(row.get("Gross earnings",0) or 0),
                    "service_fee":  float(row.get("Service fee",0) or 0),
                    "currency":     row.get("Currency","MYR").strip(),
                })
            except Exception:
                pass
    return reservations

def reconcile_airbnb_vs_manager(airbnb_rows: list, db_bookings: list) -> dict:
    """
    Compare Airbnb CSV bookings vs manager DB bookings for same month.
    Returns dict with matched, direct/extras, and summary.
    """
    # Build lookup of DB bookings by guest name (normalised)
    def norm(s): return (s or "").upper().strip().split()[0]  # first word match

    db_lookup = {}
    for b in db_bookings:
        key = norm(b.get("guest_name",""))
        if key not in db_lookup:
            db_lookup[key] = []
        db_lookup[key].append(b)

    airbnb_names = {norm(r["guest"]) for r in airbnb_rows}
    db_names     = set(db_lookup.keys())

    matched  = []
    directs  = []

    # Airbnb guests matched in DB
    for r in airbnb_rows:
        key = norm(r["guest"])
        if key in db_names:
            db_match = db_lookup[key][0]
            matched.append({
                "guest":        r["guest"],
                "checkin":      r["checkin"],
                "checkout":     r["checkout"],
                "nights":       r["nights"],
                "airbnb_payout":r["payout"],
                "airbnb_gross": r["gross"],
                "db_amount":    float(db_match.get("amount",0) or 0),
                "confirmation": r["confirmation"],
                "amount_diff":  float(db_match.get("amount",0) or 0) - r["payout"],
            })
        # else: in Airbnb but not in DB — cross-month (checkout next month)

    # DB bookings NOT in Airbnb CSV = Direct/Extras
    for b in db_bookings:
        key = norm(b.get("guest_name",""))
        if key not in airbnb_names:
            directs.append({
                "guest":    b.get("guest_name",""),
                "checkin":  b.get("checkin",""),
                "checkout": b.get("checkout",""),
                "nights":   int(b.get("nights",0) or 0),
                "amount":   float(b.get("amount",0) or 0),
                "source":   b.get("source",""),
                "db_id":    b.get("id"),
            })

    airbnb_total = sum(r["payout"] for r in airbnb_rows)
    db_total     = sum(float(b.get("amount",0) or 0) for b in db_bookings)
    direct_total = sum(d["amount"] for d in directs)

    return {
        "matched":       matched,
        "directs":       directs,
        "airbnb_total":  airbnb_total,
        "db_total":      db_total,
        "direct_total":  direct_total,
        "difference":    db_total - airbnb_total,
    }

with tab10:
    st.markdown('<div class="card">', unsafe_allow_html=True)
    st.markdown('<p class="card-title">💰 Direct Income & Extras</p>', unsafe_allow_html=True)
    st.markdown(
        "Upload the Airbnb monthly CSV — the app automatically compares it against "
        "the Manager's report already in the database and identifies Direct bookings "
        "and Extra charges. **Supplementary register only — does not affect P&L.**"
    )

    di_subtab1, di_subtab2 = st.tabs([
        "📋 All Records", "📊 Dashboard"
    ])

    # ── ALL RECORDS ─────────────────────────────────────────────────────────────
    with di_subtab1:
        rf1, rf2, rf3 = st.columns(3)
        di_filter_yr   = rf1.selectbox("Year",  list(range(datetime.datetime.now().year,2023,-1)), key="di_fyr")
        di_filter_mo   = rf2.selectbox("Month", ["All"] + MONTHS, key="di_fmo")
        di_filter_type = rf3.selectbox("Type",  ["All"] + DIRECT_INCOME_TYPES, key="di_ftype")

        conn = get_db()
        if di_filter_mo == "All":
            di_rows = conn.execute(
                "SELECT * FROM direct_income WHERE year=? ORDER BY month DESC, id DESC",
                (di_filter_yr,)
            ).fetchall()
        else:
            mo_num = MONTHS.index(di_filter_mo) + 1
            di_rows = conn.execute(
                "SELECT * FROM direct_income WHERE year=? AND month=? ORDER BY id DESC",
                (di_filter_yr, mo_num)
            ).fetchall()
        conn.close()

        if di_filter_type != "All":
            di_rows = [r for r in di_rows if r.get("income_type","") == di_filter_type]

        if not di_rows:
            st.info("No direct income entries for this period.")
        else:
            total_di = sum(float(r.get("amount",0) or 0) for r in di_rows)
            m1, m2 = st.columns(2)
            m1.metric("Total Entries", len(di_rows))
            m2.metric("Total Amount", fmt_myr(total_di))
            st.divider()

            for r in [dict(r) for r in di_rows]:
                r_id = r.get("id")
                mo_label = MONTHS[int(r.get("month",1))-1]
                with st.expander(
                    f"**{mo_label} {r.get('year')}** — {r.get('guest_name','—')} — "
                    f"{r.get('income_type','—')} — {fmt_myr(r.get('amount',0))}",
                    expanded=False
                ):
                    d1, d2, d3 = st.columns(3)
                    d1.markdown(f"**Date:** {r.get('date_received','—')}")
                    d2.markdown(f"**Method:** {r.get('payment_method','—')}")
                    d3.markdown(f"**Ref:** {r.get('reference','—')}")
                    if r.get("notes"):
                        st.markdown(f"**Notes:** {r.get('notes','')}")

                    st.divider()
                    _, del_col = st.columns([4, 1])
                    if del_col.button("🗑️ Delete", key=f"di_del_{r_id}"):
                        st.session_state[f"di_confirm_del_{r_id}"] = True

                    if st.session_state.get(f"di_confirm_del_{r_id}"):
                        st.error("Delete this record permanently?")
                        cc1, cc2 = st.columns(2)
                        if cc1.button("✅ Yes", key=f"di_del_yes_{r_id}", type="primary"):
                            conn = get_db()
                            conn.execute("DELETE FROM direct_income WHERE id=?", (r_id,))
                            conn.commit(); conn.close()
                            st.session_state.pop(f"di_confirm_del_{r_id}", None)
                            st.rerun()
                        if cc2.button("❌ Cancel", key=f"di_del_no_{r_id}"):
                            st.session_state.pop(f"di_confirm_del_{r_id}", None)
                            st.rerun()

    # ── DASHBOARD ───────────────────────────────────────────────────────────────
    with di_subtab2:
        conn = get_db()
        all_di = conn.execute(
            "SELECT * FROM direct_income ORDER BY year, month", ()
        ).fetchall()
        # Fetch ALL bookings — group in Python to avoid adapter GROUP BY issues
        all_bk_rows = conn.execute(
            "SELECT year, month, source, amount FROM bookings", ()
        ).fetchall()
        conn.close()

        if not all_di:
            st.info("No direct income entries yet. Upload an Airbnb CSV via Tab 2 → Airbnb CSV & Reconcile.")
        else:
            # Build monthly lookups in Python
            monthly_di = {}
            for r in all_di:
                key = (int(r.get("year",0)), int(r.get("month",0)))
                monthly_di[key] = monthly_di.get(key, 0) + float(r.get("amount",0) or 0)

            monthly_bk_airbnb = {}  # Airbnb-only per month
            monthly_bk_all    = {}  # All bookings per month
            for r in all_bk_rows:
                key = (int(r.get("year",0)), int(r.get("month",0)))
                amt = float(r.get("amount",0) or 0)
                monthly_bk_all[key] = monthly_bk_all.get(key, 0) + amt
                if (r.get("source","") or "").upper() == "AIRBNB":
                    monthly_bk_airbnb[key] = monthly_bk_airbnb.get(key, 0) + amt

            total_direct   = sum(monthly_di.values())
            total_airbnb   = sum(monthly_bk_airbnb.values())
            total_combined = sum(monthly_bk_all.values())

            at1, at2, at3, at4 = st.columns(4)
            at1.metric("Airbnb Income",  fmt_myr(total_airbnb))
            at2.metric("Direct/Extras",  fmt_myr(total_direct))
            at3.metric("Combined Total", fmt_myr(total_combined))
            at4.metric("Direct %",
                f"{total_direct/total_combined*100:.1f}%"
                if total_combined > 0 else "—")

            st.divider()

            # By type breakdown
            st.markdown("**Breakdown by Income Type:**")
            type_totals = {}
            for r in all_di:
                t = r.get("income_type","Other")
                type_totals[t] = type_totals.get(t, 0) + float(r.get("amount",0) or 0)
            type_rows = [
                {"Type": k, "Amount (RM)": fmt_myr(v),
                 "% of Direct": f"{v/total_direct*100:.1f}%"}
                for k,v in sorted(type_totals.items(), key=lambda x:-x[1])
            ]
            st.dataframe(pd.DataFrame(type_rows), use_container_width=True, hide_index=True)

            st.divider()

            # Monthly trend
            st.markdown("**Monthly Summary — Airbnb vs Direct/Extras:**")
            trend_rows = []
            for (yr, mo), direct_amt in sorted(monthly_di.items()):
                airbnb_mo   = monthly_bk_airbnb.get((yr, mo), 0)
                combined_mo = monthly_bk_all.get((yr, mo), 0)
                pct = direct_amt/combined_mo*100 if combined_mo > 0 else 0
                trend_rows.append({
                    "Month":              f"{MONTHS[mo-1]} {yr}",
                    "Airbnb (RM)":        fmt_myr(airbnb_mo),
                    "Direct/Extras (RM)": fmt_myr(direct_amt),
                    "Combined (RM)":      fmt_myr(combined_mo),
                    "Direct %":           f"{pct:.1f}%",
                })
            if trend_rows:
                st.dataframe(pd.DataFrame(trend_rows), use_container_width=True, hide_index=True)

            st.info(
                "💡 **Note for ULU 2 feasibility:** Use Airbnb Income only as the benchmark. "
                "Direct/Extras are non-recurring and should not be projected as future income."
            )

    st.markdown('</div>', unsafe_allow_html=True)
